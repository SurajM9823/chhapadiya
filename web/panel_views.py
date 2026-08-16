from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from .models import SiteSettings, CarouselSlide, Reel, Category, SubCategory, Country, CustomerTier, DeliveryTimeTier, Customer, Product, ProductImage, ProductTierPrice, Stat, TrustedClient, Testimonial, TeamMember, Founder, Service, WhyChooseUs, StockEntry, ContactInquiry, Order, OrderItem, CustomerUser, Role, Permission, Billing, BillingItem, Package, PackageItem, PackageImage, AboutContent, ProductAlliance, OrderPayment, ProductUnit
from .email_utils import send_order_status_update_email
from .permissions import permission_required, check_permission
import json
from django.db.models import Q
from django.http import JsonResponse
from django.views.decorators.http import require_GET
import random
import string
from datetime import datetime, timedelta
from django.core.mail import send_mail

def is_staff_user(user):
    return user.is_superuser or user.is_staff

def panel_login(request):
    if request.user.is_authenticated:
        if request.user.is_superuser or request.user.role:
            return redirect('panel_dashboard')
        else:
            messages.error(request, f'Your account does not have a role assigned. Please contact administrator.')
            logout(request)
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '').strip()
        user = authenticate(request, username=username, password=password)
        if user:
            if user.is_superuser or user.role:
                login(request, user)
                return redirect('panel_dashboard')
            else:
                messages.error(request, f'User "{username}" does not have a role assigned. Contact administrator.')
        else:
            messages.error(request, 'Invalid username or password.')
    return render(request, 'panel/login.html')

def panel_logout(request):
    logout(request)
    return redirect('panel_login')


def panel_forgot_password(request):
    if request.method == 'POST':
        email = request.POST.get('email', '').strip()
        user = CustomerUser.objects.filter(email=email).first()
        
        if user:
            # Generate 6-digit OTP
            otp = ''.join(random.choices(string.digits, k=6))
            
            # Store OTP in session with expiry (10 minutes)
            request.session['reset_otp'] = otp
            request.session['reset_email'] = email
            request.session['otp_created_at'] = datetime.now().timestamp()
            
            # Send OTP via email
            try:
                settings = SiteSettings.get()
                from_email = f'{settings.business_name} <{settings.email or "suraj20001123@gmail.com"}>'
                
                html_message = f"""
                <html>
                <head>
                    <style>
                        body {{ font-family: Arial, sans-serif; color: #333; }}
                        .container {{ max-width: 600px; margin: 0 auto; padding: 20px; }}
                        .header {{ background-color: #059669; color: white; padding: 20px; text-align: center; border-radius: 8px 8px 0 0; }}
                        .content {{ background-color: #f9fafb; padding: 30px; }}
                        .otp-box {{ background-color: white; padding: 30px; border-radius: 8px; text-align: center; border: 2px dashed #059669; }}
                        .otp-code {{ font-size: 32px; font-weight: bold; color: #059669; letter-spacing: 8px; }}
                        .footer {{ text-align: center; padding: 20px; color: #6b7280; font-size: 12px; }}
                    </style>
                </head>
                <body>
                    <div class="container">
                        <div class="header">
                            <h1 style="margin: 0;">Password Reset Request</h1>
                        </div>
                        <div class="content">
                            <p>Hello,</p>
                            <p>You requested to reset your password. Use the OTP below to verify your identity:</p>
                            <div class="otp-box">
                                <p style="margin: 0 0 10px 0; color: #6b7280; font-size: 14px;">Your OTP Code</p>
                                <div class="otp-code">{otp}</div>
                                <p style="margin: 15px 0 0 0; color: #ef4444; font-size: 12px;">This OTP will expire in 10 minutes</p>
                            </div>
                            <p style="margin-top: 20px;">If you didn't request this password reset, please ignore this email.</p>
                        </div>
                        <div class="footer">
                            <p>© {settings.business_name}. All rights reserved.</p>
                        </div>
                    </div>
                </body>
                </html>
                """
                
                send_mail(
                    subject='Password Reset OTP',
                    message=f'Your password reset OTP is: {otp}. Valid for 10 minutes.',
                    from_email=from_email,
                    recipient_list=[email],
                    html_message=html_message,
                    fail_silently=False,
                )
                
                return JsonResponse({'success': True, 'message': f'OTP has been sent to {email}'})
            except Exception as e:
                return JsonResponse({'success': False, 'error': f'Error sending email: {str(e)}'})
        else:
            return JsonResponse({'success': False, 'error': 'Email address not found.'})
    
    return render(request, 'panel/forgot_password.html')


def panel_verify_otp(request):
    if 'reset_email' not in request.session:
        return JsonResponse({'success': False, 'error': 'Session expired. Please start again.'})
    
    email = request.session.get('reset_email')
    
    if request.method == 'POST':
        entered_otp = request.POST.get('otp', '').strip()
        stored_otp = request.session.get('reset_otp')
        otp_created_at = request.session.get('otp_created_at')
        
        # Check if OTP expired (10 minutes)
        if otp_created_at:
            created_time = datetime.fromtimestamp(otp_created_at)
            if datetime.now() - created_time > timedelta(minutes=10):
                return JsonResponse({'success': False, 'error': 'OTP has expired. Please request a new one.'})
        
        if entered_otp == stored_otp:
            request.session['otp_verified'] = True
            return JsonResponse({'success': True})
        else:
            return JsonResponse({'success': False, 'error': 'Invalid OTP. Please try again.'})
    
    return render(request, 'panel/verify_otp.html', {'email': email})


def panel_reset_password(request):
    if not request.session.get('otp_verified'):
        return JsonResponse({'success': False, 'error': 'Please verify OTP first.'})
    
    email = request.session.get('reset_email')
    
    if request.method == 'POST':
        password1 = request.POST.get('password1', '').strip()
        password2 = request.POST.get('password2', '').strip()
        
        if password1 != password2:
            return JsonResponse({'success': False, 'error': 'Passwords do not match.'})
        elif len(password1) < 8:
            return JsonResponse({'success': False, 'error': 'Password must be at least 8 characters long.'})
        else:
            user = CustomerUser.objects.filter(email=email).first()
            if user:
                user.set_password(password1)
                user.save()
                
                # Clear session data
                request.session.pop('reset_otp', None)
                request.session.pop('reset_email', None)
                request.session.pop('otp_created_at', None)
                request.session.pop('otp_verified', None)
                
                return JsonResponse({'success': True})
            else:
                return JsonResponse({'success': False, 'error': 'User not found.'})
    
    return render(request, 'panel/reset_password.html')

@login_required(login_url='panel_login')
def panel_dashboard(request):
    # Allow all authenticated users with a role to see dashboard
    if not (request.user.is_superuser or request.user.role):
        messages.error(request, 'Access denied. Role assignment required.')
        return redirect('panel_login')
    from django.db.models import Sum, Count, Q
    from datetime import timedelta
    from django.utils import timezone
    
    # Check permissions for different modules
    can_view_orders = request.user.is_superuser or check_permission(request.user, 'orders', 'view')
    can_view_customers = request.user.is_superuser or check_permission(request.user, 'customers', 'view')
    can_view_products = request.user.is_superuser or check_permission(request.user, 'products', 'view')
    can_view_content = request.user.is_superuser or check_permission(request.user, 'content', 'view')
    
    # Orders Statistics (only if user has permission)
    total_orders = 0
    pending_orders = 0
    delivered_orders = 0
    cancelled_orders = 0
    delivered_income = 0
    pending_income = 0
    total_income = 0
    pending_order_list = []
    sales_by_status = []
    
    if can_view_orders:
        total_orders = Order.objects.count()
        pending_orders = Order.objects.filter(status__in=['pending', 'confirmed', 'processing']).count()
        delivered_orders = Order.objects.filter(status='delivered').count()
        cancelled_orders = Order.objects.filter(status='cancelled').count()
        delivered_income = Order.objects.filter(status='delivered').aggregate(total=Sum('total'))['total'] or 0
        pending_income = Order.objects.filter(status__in=['pending', 'confirmed', 'processing']).aggregate(total=Sum('total'))['total'] or 0
        total_income = delivered_income
        pending_order_list = Order.objects.filter(status__in=['pending', 'confirmed', 'processing']).select_related('user').prefetch_related('items')[:5]
        sales_by_status = Order.objects.values('status').annotate(count=Count('id'), total=Sum('total')).order_by('-count')
    
    # Top Customers (only if user has permission)
    top_customers_list = []
    if can_view_customers and can_view_orders:
        from django.db.models import F
        top_customers = Order.objects.filter(status='delivered').values('user__email', 'user__first_name', 'user__last_name').annotate(
            total_spent=Sum('total'),
            order_count=Count('id')
        ).order_by('-total_spent')[:5]
        
        for cust in top_customers:
            top_customers_list.append({
                'email': cust['user__email'],
                'full_name': f"{cust['user__first_name']} {cust['user__last_name']}".strip() or cust['user__email'],
                'total_spent': cust['total_spent'],
                'order_count': cust['order_count']
            })
    
    # Low Stock Products (only if user has permission)
    low_stock_products = []
    total_products = 0
    low_stock_count = 0
    out_of_stock_count = 0
    
    if can_view_products:
        total_products = Product.objects.count()
        for product in Product.objects.all():
            stock = product.stock_quantity
            if stock < 10:
                if len(low_stock_products) < 5:
                    low_stock_products.append({
                        'name': product.name,
                        'sku': product.sku,
                        'stock': stock,
                        'mrp': product.mrp
                    })
            if stock < 10:
                low_stock_count += 1
            if stock == 0:
                out_of_stock_count += 1
    
    # Recent Inquiries (only if user has permission)
    recent_inquiries = []
    total_inquiries = 0
    unread_inquiries = 0
    
    if can_view_content:
        recent_inquiries = ContactInquiry.objects.all()[:5]
        total_inquiries = ContactInquiry.objects.count()
        unread_inquiries = ContactInquiry.objects.filter(is_read=False).count()
    
    # Visitor Data (mock data - available to all)
    visitor_data = []
    for i in range(6, -1, -1):
        date = timezone.now() - timedelta(days=i)
        visitor_data.append({
            'date': date.strftime('%a'),
            'count': (i * 15 + 20) if i > 0 else 45
        })
    
    # Site Visitors (mock data - available to all)
    total_visitors = 1250
    mobile_users = 750
    desktop_users = 500
    
    context = {
        'can_view_orders': can_view_orders,
        'can_view_customers': can_view_customers,
        'can_view_products': can_view_products,
        'can_view_content': can_view_content,
        'total_orders': total_orders,
        'pending_orders': pending_orders,
        'delivered_orders': delivered_orders,
        'cancelled_orders': cancelled_orders,
        'total_income': total_income,
        'pending_income': pending_income,
        'pending_order_list': pending_order_list,
        'top_customers': top_customers_list,
        'low_stock_products': low_stock_products,
        'recent_inquiries': recent_inquiries,
        'visitor_data': visitor_data,
        'sales_by_status': sales_by_status,
        'total_visitors': total_visitors,
        'mobile_users': mobile_users,
        'desktop_users': desktop_users,
        'total_inquiries': total_inquiries,
        'unread_inquiries': unread_inquiries,
        'total_products': total_products,
        'low_stock_count': low_stock_count,
        'out_of_stock_count': out_of_stock_count,
    }
    return render(request, 'panel/dashboard.html', context)




@login_required(login_url='panel_login')
@permission_required('settings', 'view')
def panel_settings(request):
    s = SiteSettings.get()
    
    # Check if user can edit when trying to save
    if request.method == 'POST':
        if not (request.user.is_superuser or check_permission(request.user, 'settings', 'edit')):
            messages.error(request, 'You do not have permission to edit settings.')
            return redirect('panel_settings')
        
        s.business_name = request.POST.get('business_name', '')
        s.tagline = request.POST.get('tagline', '')
        s.email = request.POST.get('email', '')
        s.phone = request.POST.get('phone', '')
        s.phone2 = request.POST.get('phone2', '')
        s.address = request.POST.get('address', '')
        s.facebook = request.POST.get('facebook', '')
        s.instagram = request.POST.get('instagram', '')
        s.youtube = request.POST.get('youtube', '')
        s.whatsapp = request.POST.get('whatsapp', '')
        s.tiktok = request.POST.get('tiktok', '')
        s.linkedin = request.POST.get('linkedin', '')
        s.twitter = request.POST.get('twitter', '')
        s.map_embed = request.POST.get('map_embed', '')
        s.hours_weekday = request.POST.get('hours_weekday', '')
        s.hours_saturday = request.POST.get('hours_saturday', '')
        s.bank_name = request.POST.get('bank_name', '')
        s.bank_account_name = request.POST.get('bank_account_name', '')
        s.bank_account_number = request.POST.get('bank_account_number', '')
        s.bank_branch = request.POST.get('bank_branch', '')
        if 'logo' in request.FILES:
            s.logo = request.FILES['logo']
        if 'bank_qr' in request.FILES:
            s.bank_qr = request.FILES['bank_qr']
        s.save()
        messages.success(request, 'Settings saved.')
        return redirect('panel_settings')
    
    # Pass permission info to template
    can_edit = request.user.is_superuser or check_permission(request.user, 'settings', 'edit')
    return render(request, 'panel/settings.html', {'settings': s, 'can_edit': can_edit})


#   Carousel  

@login_required(login_url='panel_login')
@permission_required('content', 'view')
def panel_carousel(request):
    slides = CarouselSlide.objects.all().order_by('order')
    
    can_create = request.user.is_superuser or check_permission(request.user, 'content', 'create')
    can_edit = request.user.is_superuser or check_permission(request.user, 'content', 'edit')
    can_delete = request.user.is_superuser or check_permission(request.user, 'content', 'delete')
    
    return render(request, 'panel/carousel.html', {
        'slides': slides,
        'can_create': can_create,
        'can_edit': can_edit,
        'can_delete': can_delete,
    })

@login_required(login_url='panel_login')
@permission_required('content', 'view')
def panel_carousel_add(request):
    # Check create permission
    if not (request.user.is_superuser or check_permission(request.user, 'content', 'create')):
        messages.error(request, 'You do not have permission to create carousel slides.')
        return redirect('panel_carousel')
    if request.method == 'POST':
        CarouselSlide.objects.create(
            title=request.POST['title'], image=request.FILES['image'],
            order=request.POST.get('order', 0), is_active=request.POST.get('is_active') == 'on',
        )
        messages.success(request, 'Slide added.')
        return redirect('panel_carousel')
    return render(request, 'panel/carousel_form.html', {'action': 'Add'})

@login_required(login_url='panel_login')
@permission_required('content', 'view')
def panel_carousel_edit(request, pk):
    slide = get_object_or_404(CarouselSlide, pk=pk)
    
    # Check edit permission when saving
    if request.method == 'POST':
        if not (request.user.is_superuser or check_permission(request.user, 'content', 'edit')):
            messages.error(request, 'You do not have permission to edit carousel slides.')
            return redirect('panel_carousel')
        slide.title = request.POST['title']
        slide.order = request.POST.get('order', 0)
        slide.is_active = request.POST.get('is_active') == 'on'
        if 'image' in request.FILES:
            slide.image = request.FILES['image']
        slide.save()
        messages.success(request, 'Slide updated.')
        return redirect('panel_carousel')
    
    can_edit = request.user.is_superuser or check_permission(request.user, 'content', 'edit')
    return render(request, 'panel/carousel_form.html', {'action': 'Edit', 'slide': slide, 'can_edit': can_edit})

@login_required(login_url='panel_login')
@permission_required('content', 'view')
def panel_carousel_delete(request, pk):
    # Check delete permission
    if not (request.user.is_superuser or check_permission(request.user, 'content', 'delete')):
        messages.error(request, 'You do not have permission to delete carousel slides.')
        return redirect('panel_carousel')
    
    get_object_or_404(CarouselSlide, pk=pk).delete()
    messages.success(request, 'Slide deleted.')
    return redirect('panel_carousel')


#   Reels  

@login_required(login_url='panel_login')
@permission_required('content', 'view')
def panel_reels(request):
    from django.core.paginator import Paginator
    reels_list = Reel.objects.all()
    paginator = Paginator(reels_list, 10)
    page_number = request.GET.get('page', 1)
    reels = paginator.get_page(page_number)
    
    can_create = request.user.is_superuser or check_permission(request.user, 'content', 'create')
    can_edit = request.user.is_superuser or check_permission(request.user, 'content', 'edit')
    can_delete = request.user.is_superuser or check_permission(request.user, 'content', 'delete')
    
    return render(request, 'panel/reels.html', {
        'reels': reels,
        'can_create': can_create,
        'can_edit': can_edit,
        'can_delete': can_delete,
    })

@login_required(login_url='panel_login')
@permission_required('content', 'view')
def panel_reel_add(request):
    # Check create permission
    if not (request.user.is_superuser or check_permission(request.user, 'content', 'create')):
        messages.error(request, 'You do not have permission to create reels.')
        return redirect('panel_reels')
    if request.method == 'POST':
        video_type = request.POST.get('video_type', 'upload')
        data = {
            'title': request.POST['title'],
            'video_type': video_type,
            'order': request.POST.get('order', 0),
            'is_active': request.POST.get('is_active') == 'on',
        }
        if video_type == 'upload' and 'video' in request.FILES:
            data['video'] = request.FILES['video']
        elif video_type == 'youtube':
            data['youtube_url'] = request.POST.get('youtube_url', '')
        elif video_type == 'tiktok':
            data['tiktok_url'] = request.POST.get('tiktok_url', '')
        
        if 'thumbnail' in request.FILES:
            data['thumbnail'] = request.FILES['thumbnail']
        
        Reel.objects.create(**data)
        messages.success(request, 'Reel added.')
        return redirect('panel_reels')
    return render(request, 'panel/reel_form.html', {'action': 'Add'})

@login_required(login_url='panel_login')
@permission_required('content', 'view')
def panel_reel_edit(request, pk):
    reel = get_object_or_404(Reel, pk=pk)
    
    # Check edit permission when saving
    if request.method == 'POST':
        if not (request.user.is_superuser or check_permission(request.user, 'content', 'edit')):
            messages.error(request, 'You do not have permission to edit reels.')
            return redirect('panel_reels')
        reel.title = request.POST['title']
        reel.video_type = request.POST.get('video_type', 'upload')
        reel.order = request.POST.get('order', 0)
        reel.is_active = request.POST.get('is_active') == 'on'
        
        if reel.video_type == 'upload' and 'video' in request.FILES:
            reel.video = request.FILES['video']
        elif reel.video_type == 'youtube':
            reel.youtube_url = request.POST.get('youtube_url', '')
            reel.video = None
        elif reel.video_type == 'tiktok':
            reel.tiktok_url = request.POST.get('tiktok_url', '')
            reel.video = None
        
        if 'thumbnail' in request.FILES:
            reel.thumbnail = request.FILES['thumbnail']
        
        reel.save()
        messages.success(request, 'Reel updated.')
        return redirect('panel_reels')
    
    can_edit = request.user.is_superuser or check_permission(request.user, 'content', 'edit')
    return render(request, 'panel/reel_form.html', {'action': 'Edit', 'reel': reel, 'can_edit': can_edit})

@login_required(login_url='panel_login')
@permission_required('content', 'view')
def panel_reel_delete(request, pk):
    # Check delete permission
    if not (request.user.is_superuser or check_permission(request.user, 'content', 'delete')):
        messages.error(request, 'You do not have permission to delete reels.')
        return redirect('panel_reels')
    
    get_object_or_404(Reel, pk=pk).delete()
    messages.success(request, 'Reel deleted.')
    return redirect('panel_reels')


#   Categories  

@login_required(login_url='panel_login')
@permission_required('content', 'view')
def panel_categories(request):
    categories = Category.objects.prefetch_related('subcategories').all()
    
    can_create = request.user.is_superuser or check_permission(request.user, 'content', 'create')
    can_edit = request.user.is_superuser or check_permission(request.user, 'content', 'edit')
    can_delete = request.user.is_superuser or check_permission(request.user, 'content', 'delete')
    
    return render(request, 'panel/categories.html', {
        'categories': categories,
        'can_create': can_create,
        'can_edit': can_edit,
        'can_delete': can_delete,
    })

@login_required(login_url='panel_login')
@permission_required('content', 'view')
def panel_category_add(request):
    # Check create permission
    if not (request.user.is_superuser or check_permission(request.user, 'content', 'create')):
        messages.error(request, 'You do not have permission to create categories.')
        return redirect('panel_categories')
    if request.method == 'POST':
        Category.objects.create(
            name=request.POST['name'], icon=request.FILES['icon'],
            link=request.POST.get('link', '/'), order=request.POST.get('order', 0),
            is_active=request.POST.get('is_active') == 'on',
        )
        messages.success(request, 'Category added.')
        return redirect('panel_categories')
    return render(request, 'panel/category_form.html', {'action': 'Add'})

@login_required(login_url='panel_login')
@permission_required('content', 'view')
def panel_category_edit(request, pk):
    category = get_object_or_404(Category, pk=pk)
    
    # Check edit permission when saving
    if request.method == 'POST':
        if not (request.user.is_superuser or check_permission(request.user, 'content', 'edit')):
            messages.error(request, 'You do not have permission to edit categories.')
            return redirect('panel_categories')
        category.name = request.POST['name']
        category.link = request.POST.get('link', '/')
        category.order = request.POST.get('order', 0)
        category.is_active = request.POST.get('is_active') == 'on'
        if 'icon' in request.FILES:
            category.icon = request.FILES['icon']
        category.save()
        messages.success(request, 'Category updated.')
        return redirect('panel_categories')
    
    can_edit = request.user.is_superuser or check_permission(request.user, 'content', 'edit')
    return render(request, 'panel/category_form.html', {'action': 'Edit', 'category': category, 'can_edit': can_edit})

@login_required(login_url='panel_login')
@permission_required('content', 'view')
def panel_category_delete(request, pk):
    # Check delete permission
    if not (request.user.is_superuser or check_permission(request.user, 'content', 'delete')):
        messages.error(request, 'You do not have permission to delete categories.')
        return redirect('panel_categories')
    
    get_object_or_404(Category, pk=pk).delete()
    messages.success(request, 'Category deleted.')
    return redirect('panel_categories')


#   Sub Categories  

@login_required(login_url='panel_login')
@permission_required('content', 'view')
def panel_subcategories(request, cat_pk):
    category = get_object_or_404(Category, pk=cat_pk)
    return render(request, 'panel/subcategories.html', {'category': category, 'subs': category.subcategories.all()})

@login_required(login_url='panel_login')
@permission_required('content', 'create')
def panel_subcategory_add(request, cat_pk):
    category = get_object_or_404(Category, pk=cat_pk)
    if request.method == 'POST':
        SubCategory.objects.create(
            category=category, name=request.POST['name'],
            order=request.POST.get('order', 0),
            is_active=request.POST.get('is_active') == 'on',
        )
        messages.success(request, 'Sub-category added.')
        return redirect('panel_subcategories', cat_pk=cat_pk)
    return render(request, 'panel/subcategory_form.html', {'action': 'Add', 'category': category})

@login_required(login_url='panel_login')
@permission_required('content', 'edit')
def panel_subcategory_edit(request, cat_pk, pk):
    category = get_object_or_404(Category, pk=cat_pk)
    sub = get_object_or_404(SubCategory, pk=pk, category=category)
    if request.method == 'POST':
        sub.name = request.POST['name']
        sub.order = request.POST.get('order', 0)
        sub.is_active = request.POST.get('is_active') == 'on'
        sub.save()
        messages.success(request, 'Sub-category updated.')
        return redirect('panel_subcategories', cat_pk=cat_pk)
    return render(request, 'panel/subcategory_form.html', {'action': 'Edit', 'category': category, 'sub': sub})

@login_required(login_url='panel_login')
@permission_required('content', 'delete')
def panel_subcategory_delete(request, cat_pk, pk):
    get_object_or_404(SubCategory, pk=pk, category_id=cat_pk).delete()
    messages.success(request, 'Sub-category deleted.')
    return redirect('panel_subcategories', cat_pk=cat_pk)


#   Product Detail  

@login_required(login_url='panel_login')
@permission_required('products', 'view')
def panel_product_detail(request, pk):
    product = get_object_or_404(Product.objects.select_related('category', 'sub_category', 'linked_package').prefetch_related('images', 'tier_prices__tier', 'related_products', 'linked_package__items__product'), pk=pk)
    return render(request, 'panel/product_detail.html', {'product': product})


#   Countries  

@login_required(login_url='panel_login')
@permission_required('products', 'view')
def panel_countries(request):
    return render(request, 'panel/countries.html', {'countries': Country.objects.all()})

@login_required(login_url='panel_login')
@permission_required('products', 'edit')
def panel_country_save(request, pk=None):
    country = get_object_or_404(Country, pk=pk) if pk else None
    if request.method == 'POST':
        name = request.POST['name']
        code = request.POST.get('code', '')
        if country:
            country.name = name; country.code = code; country.save()
            messages.success(request, 'Country updated.')
        else:
            Country.objects.create(name=name, code=code)
            messages.success(request, 'Country added.')
        return redirect('panel_countries')
    return render(request, 'panel/country_form.html', {'action': 'Edit' if country else 'Add', 'country': country})

@login_required(login_url='panel_login')
@permission_required('products', 'delete')
def panel_country_delete(request, pk):
    get_object_or_404(Country, pk=pk).delete()
    messages.success(request, 'Country deleted.')
    return redirect('panel_countries')


#   Customer Tiers  

@login_required(login_url='panel_login')
@permission_required('customers', 'view')
def panel_tiers(request):
    can_create = request.user.is_superuser or check_permission(request.user, 'customers', 'create')
    can_edit = request.user.is_superuser or check_permission(request.user, 'customers', 'edit')
    can_delete = request.user.is_superuser or check_permission(request.user, 'customers', 'delete')
    
    return render(request, 'panel/tiers.html', {
        'tiers': CustomerTier.objects.all(),
        'can_create': can_create,
        'can_edit': can_edit,
        'can_delete': can_delete,
    })

@login_required(login_url='panel_login')
@permission_required('customers', 'create')
def panel_tier_add(request):
    if request.method == 'POST':
        CustomerTier.objects.create(
            name=request.POST['name'], description=request.POST.get('description', ''),
            order=request.POST.get('order', 0), is_active=request.POST.get('is_active') == 'on',
        )
        messages.success(request, 'Tier added.')
        return redirect('panel_tiers')
    return render(request, 'panel/tier_form.html', {'action': 'Add'})

@login_required(login_url='panel_login')
@permission_required('customers', 'edit')
def panel_tier_edit(request, pk):
    tier = get_object_or_404(CustomerTier, pk=pk)
    if request.method == 'POST':
        tier.name = request.POST['name']
        tier.description = request.POST.get('description', '')
        tier.order = request.POST.get('order', 0)
        tier.is_active = request.POST.get('is_active') == 'on'
        tier.save()
        messages.success(request, 'Tier updated.')
        return redirect('panel_tiers')
    return render(request, 'panel/tier_form.html', {'action': 'Edit', 'tier': tier})

@login_required(login_url='panel_login')
@permission_required('customers', 'delete')
def panel_tier_delete(request, pk):
    get_object_or_404(CustomerTier, pk=pk).delete()
    messages.success(request, 'Tier deleted.')
    return redirect('panel_tiers')


# -- Delivery Time Tiers ------------------------------------------------------

@login_required(login_url='panel_login')
@permission_required('settings', 'view')
def panel_delivery_times(request):
    can_create = request.user.is_superuser or check_permission(request.user, 'settings', 'create')
    can_edit = request.user.is_superuser or check_permission(request.user, 'settings', 'edit')
    can_delete = request.user.is_superuser or check_permission(request.user, 'settings', 'delete')
    
    return render(request, 'panel/delivery_times.html', {
        'delivery_times': DeliveryTimeTier.objects.all(),
        'can_create': can_create,
        'can_edit': can_edit,
        'can_delete': can_delete,
    })

@login_required(login_url='panel_login')
@permission_required('settings', 'create')
def panel_delivery_time_add(request):
    if request.method == 'POST':
        DeliveryTimeTier.objects.create(
            name=request.POST['name'],
            min_time=request.POST['min_time'],
            min_unit=request.POST['min_unit'],
            max_time=request.POST['max_time'],
            max_unit=request.POST['max_unit'],
            description=request.POST.get('description', ''),
            order=request.POST.get('order', 0),
            is_active=request.POST.get('is_active') == 'on',
        )
        messages.success(request, 'Delivery time tier added.')
        return redirect('panel_delivery_times')
    return render(request, 'panel/delivery_time_form.html', {'action': 'Add'})

@login_required(login_url='panel_login')
@permission_required('settings', 'edit')
def panel_delivery_time_edit(request, pk):
    delivery_time = get_object_or_404(DeliveryTimeTier, pk=pk)
    if request.method == 'POST':
        delivery_time.name = request.POST['name']
        delivery_time.min_time = request.POST['min_time']
        delivery_time.min_unit = request.POST['min_unit']
        delivery_time.max_time = request.POST['max_time']
        delivery_time.max_unit = request.POST['max_unit']
        delivery_time.description = request.POST.get('description', '')
        delivery_time.order = request.POST.get('order', 0)
        delivery_time.is_active = request.POST.get('is_active') == 'on'
        delivery_time.save()
        messages.success(request, 'Delivery time tier updated.')
        return redirect('panel_delivery_times')
    return render(request, 'panel/delivery_time_form.html', {'action': 'Edit', 'delivery_time': delivery_time})

@login_required(login_url='panel_login')
@permission_required('settings', 'delete')
def panel_delivery_time_delete(request, pk):
    get_object_or_404(DeliveryTimeTier, pk=pk).delete()
    messages.success(request, 'Delivery time tier deleted.')
    return redirect('panel_delivery_times')


#   Customers  

from django.http import JsonResponse
from django.core.paginator import Paginator
from django.db.models import Q

from django.shortcuts import render
from django.http import JsonResponse
from django.db.models import Q
from django.contrib.auth.decorators import login_required

# Main View
@login_required(login_url='panel_login')
@permission_required('customers', 'view')
def panel_customers(request):
    can_create = request.user.is_superuser or check_permission(request.user, 'customers', 'create')
    can_edit = request.user.is_superuser or check_permission(request.user, 'customers', 'edit')
    can_delete = request.user.is_superuser or check_permission(request.user, 'customers', 'delete')

    customers = Customer.objects.select_related('tier').order_by('-created_at')

    return render(request, 'panel/customers.html', {
        'customers': customers,
        'can_create': can_create,
        'can_edit': can_edit,
        'can_delete': can_delete,
    })


# AJAX Search View
@login_required(login_url='panel_login')
def panel_customers_search(request):
    query = request.GET.get('search', '').strip()
    
    customers = Customer.objects.select_related('tier').order_by('-created_at')

    if query:
        customers = customers.filter(
            Q(name__icontains=query) |
            Q(email__icontains=query) |
            Q(phone__icontains=query) |
            Q(company__icontains=query)
        )

    data = []
    for c in customers[:100]:
        data.append({
            'id': c.id,
            'name': c.name,
            'email': c.email,
            'phone': c.phone or ' ',
            'customer_type': c.get_customer_type_display() or c.customer_type.title(),
            'company': c.company or ' ',
            'tier': c.tier.name if c.tier else 'Normal',
            'is_active': c.is_active,
        })

    return JsonResponse({'customers': data})

@login_required(login_url='panel_login')
@permission_required('customers', 'create')
def panel_customer_add(request):
    tiers = CustomerTier.objects.filter(is_active=True)
    
    if request.method == 'POST':
        Customer.objects.create(
            name=request.POST['name'],
            email=request.POST['email'],
            phone=request.POST.get('phone', ''),
            pan_number=request.POST.get('pan_number', '').upper(),
            company=request.POST.get('company', ''),
            address=request.POST.get('address', ''),
            customer_type=request.POST.get('customer_type', 'retailer'),  # ? Added
            tier_id=request.POST.get('tier') or None,
            is_active=request.POST.get('is_active') == 'on',
        )
        messages.success(request, 'Customer added successfully.')
        return redirect('panel_customers')

    return render(request, 'panel/customer_form.html', {
        'action': 'Add',
        'tiers': tiers,
        'customer_types': Customer.CUSTOMER_TYPE_CHOICES,   # ? Pass choices
    })


@login_required(login_url='panel_login')
@permission_required('customers', 'edit')
def panel_customer_edit(request, pk):
    customer = get_object_or_404(Customer, pk=pk)
    tiers = CustomerTier.objects.filter(is_active=True)

    if request.method == 'POST':
        customer.name = request.POST['name']
        customer.email = request.POST['email']
        customer.phone = request.POST.get('phone', '')
        customer.pan_number = request.POST.get('pan_number', '').upper()
        customer.company = request.POST.get('company', '')
        customer.address = request.POST.get('address', '')
        customer.customer_type = request.POST.get('customer_type', 'retailer')   # ? Added
        customer.tier_id = request.POST.get('tier') or None
        customer.is_active = request.POST.get('is_active') == 'on'
        customer.save()
        
        messages.success(request, 'Customer updated successfully.')
        return redirect('panel_customers')

    return render(request, 'panel/customer_form.html', {
        'action': 'Edit',
        'customer': customer,
        'tiers': tiers,
        'customer_types': Customer.CUSTOMER_TYPE_CHOICES,   # ? Pass choices
    })

@login_required(login_url='panel_login')
@permission_required('customers', 'delete')
def panel_customer_delete(request, pk):
    get_object_or_404(Customer, pk=pk).delete()
    messages.success(request, 'Customer deleted.')
    return redirect('panel_customers')



@login_required(login_url='panel_login')
@permission_required('products', 'view')
def panel_products(request):
    from django.core.paginator import Paginator
    from django.db.models import Q
    
    search = request.GET.get('search', '').strip()
    category_id = request.GET.get('category', '').strip()
    status = request.GET.get('status', '').strip()
    page = request.GET.get('page', 1)
    limit = request.GET.get('limit', 10)
    
    try:
        limit = int(limit)
        if limit > 100:
            limit = 100
        if limit < 5:
            limit = 5
    except:
        limit = 20
    
    qs = Product.objects.select_related('category', 'sub_category').all()
    
    if search:
        qs = qs.filter(Q(name__icontains=search) | Q(sku__icontains=search) | Q(brand__icontains=search))
    
    if category_id:
        try:
            qs = qs.filter(category_id=int(category_id))
        except:
            pass
    
    if status:
        if status == 'active':
            qs = qs.filter(is_active=True)
        elif status == 'inactive':
            qs = qs.filter(is_active=False)
        elif status == 'featured':
            qs = qs.filter(is_featured=True)
    
    qs = qs.order_by('-created_at')
    
    paginator = Paginator(qs, limit)
    products = paginator.get_page(page)
    categories = Category.objects.filter(is_active=True).values('id', 'name').order_by('name')
    
    can_create = request.user.is_superuser or check_permission(request.user, 'products', 'create')
    can_edit = request.user.is_superuser or check_permission(request.user, 'products', 'edit')
    can_delete = request.user.is_superuser or check_permission(request.user, 'products', 'delete')
    can_export = request.user.is_superuser or check_permission(request.user, 'products', 'view')
    
    context = {
        'products': products,
        'categories': list(categories),
        'search': search,
        'category_id': category_id,
        'status': status,
        'limit': limit,
        'total_count': paginator.count,
        'can_create': can_create,
        'can_edit': can_edit,
        'can_delete': can_delete,
        'can_export': can_export,
    }

    # ? Return only the table fragment for AJAX requests
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return render(request, 'panel/products_fragment.html', context)
    
    return render(request, 'panel/products.html', context)


@login_required(login_url='panel_login')
@permission_required('products', 'view')
def api_product_units(request):
    """GET: list units. POST: get-or-create a unit by name."""
    if request.method == 'POST':
        import json as _json
        data = _json.loads(request.body)
        name = data.get('name', '').strip()
        if not name:
            return JsonResponse({'error': 'Name required'}, status=400)
        unit, _ = ProductUnit.objects.get_or_create(name__iexact=name, defaults={'name': name})
        return JsonResponse({'id': unit.pk, 'name': unit.name})
    q = request.GET.get('q', '').strip()
    qs = ProductUnit.objects.all()
    if q:
        qs = qs.filter(name__icontains=q)
    return JsonResponse([{'id': u.pk, 'name': u.name} for u in qs], safe=False)


@login_required(login_url='panel_login')
@permission_required('products', 'create')
def panel_product_add(request):
    categories = Category.objects.prefetch_related('subcategories').filter(is_active=True)
    tiers = CustomerTier.objects.filter(is_active=True)
    all_products = Product.objects.all()
    countries = Country.objects.all()
    delivery_times = DeliveryTimeTier.objects.filter(is_active=True)
    packages = Package.objects.filter(is_active=True)
    if request.method == 'POST':
        unit_id = request.POST.get('unit') or None
        product = Product.objects.create(
            name=request.POST['name'], sku=request.POST['sku'],
            product_code=request.POST.get('product_code', ''),
            brand=request.POST.get('brand', ''), origin_id=request.POST.get('origin') or None,
            category_id=request.POST.get('category') or None,
            sub_category_id=request.POST.get('sub_category') or None,
            short_description=request.POST.get('short_description', ''),
            full_description=request.POST.get('full_description', ''),
            specifications=request.POST.get('specifications', ''),
            mrp=request.POST['mrp'],
            retail_price=request.POST.get('retail_price') or None,
            dealer_price=request.POST.get('dealer_price') or None,
            tax_included=request.POST.get('tax_included') == 'on',
            tax_percent=request.POST.get('tax_percent') or 0,
            delivery_time_id=request.POST.get('delivery_time') or None,
            linked_package_id=request.POST.get('linked_package') or None,
            unit_id=unit_id,
            is_active=request.POST.get('is_active') == 'on',
            is_featured=request.POST.get('is_featured') == 'on',
        )
        for img in request.FILES.getlist('images'):
            ProductImage.objects.create(product=product, image=img)
        if product.images.exists():
            first = product.images.first(); first.is_primary = True; first.save()
        for tier in tiers:
            price_val = request.POST.get(f'tier_price_{tier.pk}')
            if price_val:
                ProductTierPrice.objects.create(product=product, tier=tier, price=price_val)
        related_ids = request.POST.getlist('related_products')
        if related_ids:
            product.related_products.set(related_ids)
        initial_stock = request.POST.get('initial_stock')
        if initial_stock and int(initial_stock) > 0:
            StockEntry.objects.create(product=product, entry_type='import', quantity_change=int(initial_stock), note='Initial stock')
        messages.success(request, 'Product added.')
        return redirect('panel_products')
    return render(request, 'panel/product_form.html', {
        'action': 'Add', 'categories': categories, 'tiers': tiers,
        'all_products': all_products, 'countries': countries, 'delivery_times': delivery_times, 'packages': packages,
    })

@login_required(login_url='panel_login')
@permission_required('products', 'edit')
def panel_product_edit(request, pk):
    product = get_object_or_404(Product, pk=pk)
    categories = Category.objects.prefetch_related('subcategories').filter(is_active=True)
    tiers = CustomerTier.objects.filter(is_active=True)
    all_products = Product.objects.exclude(pk=pk)
    countries = Country.objects.all()
    delivery_times = DeliveryTimeTier.objects.filter(is_active=True)
    packages = Package.objects.filter(is_active=True)
    tier_prices = {tp.tier_id: tp.price for tp in product.tier_prices.all()}
    if request.method == 'POST':
        product.name = request.POST['name']
        product.sku = request.POST['sku']
        product.product_code = request.POST.get('product_code', '')
        product.brand = request.POST.get('brand', '')
        product.origin_id = request.POST.get('origin') or None
        product.category_id = request.POST.get('category') or None
        product.sub_category_id = request.POST.get('sub_category') or None
        product.short_description = request.POST.get('short_description', '')
        product.full_description = request.POST.get('full_description', '')
        product.specifications = request.POST.get('specifications', '')
        product.mrp = request.POST['mrp']
        product.retail_price = request.POST.get('retail_price') or None
        product.dealer_price = request.POST.get('dealer_price') or None
        product.tax_included = request.POST.get('tax_included') == 'on'
        product.tax_percent = request.POST.get('tax_percent') or 0
        product.delivery_time_id = request.POST.get('delivery_time') or None
        product.linked_package_id = request.POST.get('linked_package') or None
        product.unit_id = request.POST.get('unit') or None
        product.is_active = request.POST.get('is_active') == 'on'
        product.is_featured = request.POST.get('is_featured') == 'on'
        product.save()
        for img in request.FILES.getlist('images'):
            ProductImage.objects.create(product=product, image=img)
        if not product.images.filter(is_primary=True).exists() and product.images.exists():
            first = product.images.first(); first.is_primary = True; first.save()
        for img_id in request.POST.getlist('delete_images'):
            ProductImage.objects.filter(pk=img_id, product=product).delete()
        for tier in tiers:
            price_val = request.POST.get(f'tier_price_{tier.pk}')
            if price_val:
                ProductTierPrice.objects.update_or_create(product=product, tier=tier, defaults={'price': price_val})
            else:
                ProductTierPrice.objects.filter(product=product, tier=tier).delete()
        product.related_products.set(request.POST.getlist('related_products'))
        messages.success(request, 'Product updated.')
        return redirect('panel_products')
    return render(request, 'panel/product_form.html', {
        'action': 'Edit', 'product': product, 'categories': categories,
        'tiers': tiers, 'tier_prices': tier_prices, 'all_products': all_products, 'countries': countries, 'delivery_times': delivery_times, 'packages': packages,
    })

from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
import json

@login_required(login_url='panel_login')
@permission_required('products', 'edit')
def panel_product_alliance(request, pk):
    """GET: fetch current alliances + searchable product list. POST: add. DELETE: remove."""
    product = get_object_or_404(Product, pk=pk)

    if request.method == 'GET':
        search = request.GET.get('q', '').strip()
        
        # Current alliances for this product
        alliances = ProductAlliance.objects.filter(
            product=product, is_active=True
        ).select_related('alliance_product').order_by('priority', '-created_at')
        
        alliance_data = [
            {
                'id': a.id,
                'product_id': a.alliance_product.pk,
                'name': a.alliance_product.name,
                'sku': a.alliance_product.sku,
                'discount_percent': str(a.discount_percent),
                'priority': a.priority,
            }
            for a in alliances
        ]
        
        # Searchable products (exclude self + already allied)
        allied_ids = [a['product_id'] for a in alliance_data]
        qs = Product.objects.exclude(pk=pk).exclude(pk__in=allied_ids).filter(is_active=True)
        if search:
            qs = qs.filter(
                Q(name__icontains=search) | Q(sku__icontains=search) | Q(brand__icontains=search)
            )
        qs = qs.values('id', 'name', 'sku', 'brand')[:20]
        
        return JsonResponse({
            'product': {'id': product.pk, 'name': product.name, 'sku': product.sku},
            'alliances': alliance_data,
            'available': list(qs),
        })

    elif request.method == 'POST':
        data = json.loads(request.body)
        alliance_product_id = data.get('alliance_product_id')
        discount_percent = data.get('discount_percent', 0)
        priority = data.get('priority', 0)
        
        alliance_product = get_object_or_404(Product, pk=alliance_product_id)
        
        if alliance_product.pk == product.pk:
            return JsonResponse({'error': 'Cannot ally a product with itself.'}, status=400)
        
        alliance, created = ProductAlliance.objects.get_or_create(
            product=product,
            alliance_product=alliance_product,
            defaults={
                'discount_percent': discount_percent,
                'priority': priority,
                'is_active': True,
            }
        )
        if not created:
            # Reactivate if it was soft-deleted or just update
            alliance.is_active = True
            alliance.discount_percent = discount_percent
            alliance.priority = priority
            alliance.save()
        
        return JsonResponse({
            'id': alliance.id,
            'product_id': alliance_product.pk,
            'name': alliance_product.name,
            'sku': alliance_product.sku,
            'discount_percent': str(alliance.discount_percent),
            'priority': alliance.priority,
        })

    elif request.method == 'DELETE':
        data = json.loads(request.body)
        alliance_id = data.get('alliance_id')
        ProductAlliance.objects.filter(pk=alliance_id, product=product).delete()
        return JsonResponse({'deleted': True})

    return JsonResponse({'error': 'Method not allowed'}, status=405)

@login_required(login_url='panel_login')
@permission_required('products', 'delete')
def panel_product_delete(request, pk):
    get_object_or_404(Product, pk=pk).delete()
    messages.success(request, 'Product deleted.')
    return redirect('panel_products')


@login_required(login_url='panel_login')
@permission_required('products', 'view')
def panel_products_export(request):
    from django.http import HttpResponse
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from django.db.models import Q
 
    search     = request.GET.get('search', '').strip()
    category_id = request.GET.get('category', '').strip()
    status     = request.GET.get('status', '').strip()
 
    qs = Product.objects.select_related(
        'category', 'sub_category', 'origin', 'delivery_time', 'linked_package'
    ).prefetch_related('tier_prices__tier', 'images').all()
 
    if search:
        qs = qs.filter(Q(name__icontains=search) | Q(sku__icontains=search) | Q(brand__icontains=search))
    if category_id:
        try:
            qs = qs.filter(category_id=int(category_id))
        except Exception:
            pass
    if status == 'active':
        qs = qs.filter(is_active=True)
    elif status == 'inactive':
        qs = qs.filter(is_active=False)
    elif status == 'featured':
        qs = qs.filter(is_featured=True)
 
    qs = qs.order_by('name')
 
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Products'
 
    # -- Styles --------------------------------------------------------------
    header_fill  = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    section_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
    alt_fill     = PatternFill(start_color='F7FBFF', end_color='F7FBFF', fill_type='solid')
    header_font  = Font(bold=True, color='FFFFFF', size=10)
    section_font = Font(bold=True, color='1F4E79', size=10)
    thin_border  = Border(
        left=Side(style='thin', color='BDD7EE'),
        right=Side(style='thin', color='BDD7EE'),
        top=Side(style='thin', color='BDD7EE'),
        bottom=Side(style='thin', color='BDD7EE'),
    )
    center = Alignment(horizontal='center', vertical='top', wrap_text=True)
    left   = Alignment(horizontal='left',   vertical='top', wrap_text=True)
 
    # -- Column definitions: (header, width, alignment) -------------------
    # IMPORTANT: import view reads columns in exactly this order (col A=1   )
    columns = [
        # --- Identity ---
        ('SKU *',            15, center),
        ('Product Code',     15, center),
        ('Name *',           42, left),
        ('Brand',            20, left),
        # --- Classification ---
        ('Category',         22, left),
        ('Sub-Category',     22, left),
        ('Origin Country',   18, left),
        # --- Pricing ---
        ('MRP (Rs.) *',      13, center),
        ('Retail Price',     13, center),
        ('Dealer Price',     13, center),
        ('Tax Included',     13, center),   # Yes / No
        ('Tax %',            10, center),
        # --- Stock & Logistics ---
        ('Initial Stock',    13, center),
        ('Delivery Time',    25, left),
        ('Linked Package',   25, left),
        # --- Content ---
        ('Short Description',50, left),
        ('Full Description', 60, left),
        ('Specifications',   50, left),
        # --- Flags ---
        ('Active',           10, center),   # Yes / No
        ('Featured',         10, center),   # Yes / No
        # --- Read-only reference (ignored on import) ---
        ('Stock Qty (ref)',  13, center),
        ('Primary Image URL',60, left),
        ('All Image URLs',   80, left),
        ('Tier Prices (ref)',40, left),
        ('Created At',       22, center),
        ('Updated At',       22, center),
    ]
 
    # -- Header row -------------------------------------------------------
    ws.row_dimensions[1].height = 28
    for col_num, (header, width, align) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill   = header_fill
        cell.font   = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(col_num)].width = width
 
    ws.freeze_panes = 'A2'
 
    # -- Data rows --------------------------------------------------------
    for row_num, p in enumerate(qs, 2):
        row_fill = alt_fill if row_num % 2 == 0 else None
 
        def w(col, value, align=left):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.alignment = align
            cell.border    = thin_border
            if row_fill:
                cell.fill = row_fill
 
        # Identity
        w(1,  p.sku,                                center)
        w(2,  p.product_code,                       center)
        w(3,  p.name)
        w(4,  p.brand)
        # Classification
        w(5,  p.category.name     if p.category     else '')
        w(6,  p.sub_category.name if p.sub_category else '')
        w(7,  p.origin.name       if p.origin       else '')
        # Pricing
        w(8,  float(p.mrp),                                            center)
        w(9,  float(p.retail_price) if p.retail_price else float(p.mrp), center)
        w(10, float(p.dealer_price) if p.dealer_price else float(p.mrp), center)
        w(11, 'Yes' if p.tax_included else 'No',                       center)
        w(12, float(p.tax_percent),                                    center)
        # Stock & Logistics
        w(13, 0,  center)   # Initial Stock   always 0 for existing products
        w(14, str(p.delivery_time)     if p.delivery_time  else '')
        w(15, p.linked_package.name    if p.linked_package else '')
        # Content
        w(16, p.short_description)
        w(17, p.full_description)
        w(18, p.specifications)
        # Flags
        w(19, 'Yes' if p.is_active   else 'No', center)
        w(20, 'Yes' if p.is_featured else 'No', center)
        # Read-only reference
        w(21, p.stock_quantity, center)
        primary = p.primary_image
        w(22, request.build_absolute_uri(primary.image.url) if primary else '')
        all_imgs = ', '.join(request.build_absolute_uri(i.image.url) for i in p.images.all())
        w(23, all_imgs)
        tier_txt = ', '.join(f"{tp.tier.name}: Rs.{tp.price}" for tp in p.tier_prices.all())
        w(24, tier_txt)
        w(25, p.created_at.strftime('%Y-%m-%d %H:%M'), center)
        w(26, p.updated_at.strftime('%Y-%m-%d %H:%M'), center)
 
    # Mark read-only reference columns with a different header fill
    for col in (21, 22, 23, 24, 25, 26):
        cell = ws.cell(row=1, column=col)
        cell.fill = PatternFill(start_color='7F7F7F', end_color='7F7F7F', fill_type='solid')
 
    # -- Add a legend sheet -----------------------------------------------
    ls = wb.create_sheet('How To Use')
    notes = [
        ('Column', 'Notes'),
        ('SKU *',            'Required. Must be unique. Used to match existing products on update.'),
        ('Name *',           'Required.'),
        ('MRP *',            'Required. Numeric.'),
        ('Retail / Dealer',  'Leave blank to default to MRP.'),
        ('Tax Included',     'Yes or No'),
        ('Active / Featured','Yes or No'),
        ('Initial Stock',    'Only applied on NEW products. Ignored when updating existing SKUs.'),
        ('Delivery Time',    'Must match an existing Delivery Time name exactly.'),
        ('Linked Package',   'Must match an existing Package name exactly.'),
        ('Grey columns',     'Read-only reference   ignored during import.'),
    ]
    ls.column_dimensions['A'].width = 22
    ls.column_dimensions['B'].width = 70
    for r, (a, b) in enumerate(notes, 1):
        ls.cell(r, 1, a).font = Font(bold=(r == 1))
        ls.cell(r, 2, b)
 
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=products_export.xlsx'
    wb.save(response)
    return response
 
 
# -----------------------------------------------------------------------------
#  IMPORT
# -----------------------------------------------------------------------------
@login_required(login_url='panel_login')
@permission_required('products', 'create')
def panel_products_import(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        import openpyxl
 
        excel_file = request.FILES['excel_file']
        try:
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            ws = wb.active
 
            created_count = 0
            updated_count = 0
            errors = []
 
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                # Skip entirely blank rows
                if not any(row):
                    continue
 
                def cell(n):
                    """Return stripped string for column n (1-based)."""
                    v = row[n - 1] if len(row) >= n else None
                    return str(v).strip() if v not in (None, '') else ''
 
                def yn(n):
                    return cell(n).lower() in ('yes', 'true', '1', 'y')
 
                # -- Required fields --------------------------------------
                sku  = cell(1)
                name = cell(3)
                mrp_raw = cell(8)
 
                if not sku:
                    errors.append(f"Row {row_idx}: SKU is required   row skipped.")
                    continue
                if not name:
                    errors.append(f"Row {row_idx}: Name is required (SKU: {sku})   row skipped.")
                    continue
                try:
                    mrp = float(mrp_raw)
                except (ValueError, TypeError):
                    errors.append(f"Row {row_idx}: MRP must be a number (SKU: {sku})   row skipped.")
                    continue
 
                # -- Optional numeric fields ------------------------------
                def to_float_or_none(n):
                    v = cell(n)
                    if not v:
                        return None
                    try:
                        return float(v)
                    except ValueError:
                        return None
 
                retail_price = to_float_or_none(9)
                dealer_price = to_float_or_none(10)
                tax_percent  = to_float_or_none(12) or 0
 
                try:
                    initial_stock = int(float(cell(13))) if cell(13) else 0
                except ValueError:
                    initial_stock = 0
 
                # -- FK lookups -------------------------------------------
                category = sub_category = origin = delivery_time = linked_package = None
 
                cat_name = cell(5)
                if cat_name:
                    category = Category.objects.filter(name__iexact=cat_name, is_active=True).first()
                    if not category:
                        errors.append(f"Row {row_idx}: Category '{cat_name}' not found (SKU: {sku})   skipped.")
                        continue
 
                sub_name = cell(6)
                if sub_name:
                    qs_sub = SubCategory.objects.filter(name__iexact=sub_name, is_active=True)
                    if category:
                        qs_sub = qs_sub.filter(category=category)
                    sub_category = qs_sub.first()
                    if not sub_category:
                        errors.append(f"Row {row_idx}: Sub-category '{sub_name}' not found (SKU: {sku})   skipped.")
                        continue
 
                origin_name = cell(7)
                if origin_name:
                    origin = Country.objects.filter(name__iexact=origin_name).first()
                    if not origin:
                        errors.append(f"Row {row_idx}: Country '{origin_name}' not found (SKU: {sku})   skipped.")
                        continue
 
                dt_name = cell(14)
                if dt_name:
                    delivery_time = DeliveryTimeTier.objects.filter(
                        name__iexact=dt_name, is_active=True
                    ).first()
 
                pkg_name = cell(15)
                if pkg_name:
                    linked_package = Package.objects.filter(
                        name__iexact=pkg_name, is_active=True
                    ).first()
 
                # -- Create or Update -------------------------------------
                existing = Product.objects.filter(sku=sku).first()
 
                fields = dict(
                    name             = name,
                    product_code     = cell(2),
                    brand            = cell(4),
                    category         = category,
                    sub_category     = sub_category,
                    origin           = origin,
                    mrp              = mrp,
                    retail_price     = retail_price,
                    dealer_price     = dealer_price,
                    tax_included     = yn(11),
                    tax_percent      = tax_percent,
                    delivery_time    = delivery_time,
                    linked_package   = linked_package,
                    short_description= cell(16),
                    full_description = cell(17),
                    specifications   = cell(18),
                    is_active        = yn(19),
                    is_featured      = yn(20),
                )
 
                if existing:
                    for attr, val in fields.items():
                        setattr(existing, attr, val)
                    existing.save()
                    updated_count += 1
                else:
                    product = Product.objects.create(**fields)
                    if initial_stock > 0:
                        StockEntry.objects.create(
                            product=product,
                            entry_type='import',
                            quantity_change=initial_stock,
                            note='Imported via Excel',
                        )
                    created_count += 1
 
            # -- Flash messages -------------------------------------------
            if created_count:
                messages.success(request, f'? {created_count} product(s) created.')
            if updated_count:
                messages.info(request, f'?? {updated_count} product(s) updated.')
            if not created_count and not updated_count and not errors:
                messages.warning(request, 'No data rows found in the file.')
            for err in errors[:15]:
                messages.warning(request, err)
            if len(errors) > 15:
                messages.warning(request, f'  and {len(errors) - 15} more errors. Fix and re-import.')
 
        except Exception as e:
            messages.error(request, f'Import failed: {str(e)}')
 
        return redirect('panel_products')
 
    categories = Category.objects.filter(is_active=True).prefetch_related('subcategories')
    countries  = Country.objects.all()
    tiers      = CustomerTier.objects.filter(is_active=True)
    return render(request, 'panel/products_import.html', {
        'categories': categories,
        'countries':  countries,
        'tiers':      tiers,
    })


@login_required(login_url='panel_login')
@permission_required('stock', 'edit')
def panel_stock(request, pk):
    product = get_object_or_404(Product, pk=pk)
    entries = product.stock_entries.select_related('customer').all()
    customers = Customer.objects.filter(is_active=True).select_related('tier')
    if request.method == 'POST':
        entry_type = request.POST['entry_type']
        qty = abs(int(request.POST['quantity_change']))
        if entry_type in ('sale', 'adjustment_out'):
            qty = -qty
        customer_id = request.POST.get('customer_id') or None
        if request.POST.get('new_customer_name') and not customer_id:
            new_c = Customer.objects.create(
                name=request.POST['new_customer_name'],
                email=request.POST.get('new_customer_email', f"guest_{Product.objects.count()}@guest.local"),
                phone=request.POST.get('new_customer_phone', ''),
            )
            customer_id = new_c.pk
        entry = StockEntry.objects.create(
            product=product, entry_type=entry_type, quantity_change=qty,
            note=request.POST.get('note', ''), customer_id=customer_id,
            unit_price=request.POST.get('unit_price') or None,
        )
        messages.success(request, 'Stock entry added.')
        return redirect(f"{request.path}?receipt={entry.pk}")
    receipt_entry = None
    receipt_id = request.GET.get('receipt')
    if receipt_id:
        receipt_entry = StockEntry.objects.filter(pk=receipt_id, product=product).select_related('customer__tier').first()
    return render(request, 'panel/stock.html', {
        'product': product, 'entries': entries,
        'customers': customers, 'receipt_entry': receipt_entry,
    })


#   Services  

@login_required(login_url='panel_login')
@permission_required('content', 'view')
def panel_services(request):
    can_create = request.user.is_superuser or check_permission(request.user, 'content', 'create')
    can_edit = request.user.is_superuser or check_permission(request.user, 'content', 'edit')
    can_delete = request.user.is_superuser or check_permission(request.user, 'content', 'delete')
    
    return render(request, 'panel/services.html', {
        'services': Service.objects.all(),
        'why_items': WhyChooseUs.objects.all(),
        'can_create': can_create,
        'can_edit': can_edit,
        'can_delete': can_delete,
    })

@login_required(login_url='panel_login')
@permission_required('content', 'view')
def panel_service_add(request):
    # Check create permission
    if not (request.user.is_superuser or check_permission(request.user, 'content', 'create')):
        messages.error(request, 'You do not have permission to create services.')
        return redirect('panel_services')
    if request.method == 'POST':
        Service.objects.create(
            title=request.POST['title'], description=request.POST['description'],
            icon=request.POST.get('icon', ''), order=request.POST.get('order', 0),
            is_active=request.POST.get('is_active') == 'on',
            image=request.FILES.get('image') or None,
        )
        messages.success(request, 'Service added.')
        return redirect('panel_services')
    return render(request, 'panel/service_form.html', {'action': 'Add'})

@login_required(login_url='panel_login')
@permission_required('content', 'view')
def panel_service_edit(request, pk):
    service = get_object_or_404(Service, pk=pk)
    
    # Check edit permission when saving
    if request.method == 'POST':
        if not (request.user.is_superuser or check_permission(request.user, 'content', 'edit')):
            messages.error(request, 'You do not have permission to edit services.')
            return redirect('panel_services')
        service.title = request.POST['title']
        service.description = request.POST['description']
        service.icon = request.POST.get('icon', '')
        service.order = request.POST.get('order', 0)
        service.is_active = request.POST.get('is_active') == 'on'
        if 'image' in request.FILES:
            service.image = request.FILES['image']
        service.save()
        messages.success(request, 'Service updated.')
        return redirect('panel_services')
    
    can_edit = request.user.is_superuser or check_permission(request.user, 'content', 'edit')
    return render(request, 'panel/service_form.html', {'action': 'Edit', 'service': service, 'can_edit': can_edit})

@login_required(login_url='panel_login')
@permission_required('content', 'view')
def panel_service_delete(request, pk):
    # Check delete permission
    if not (request.user.is_superuser or check_permission(request.user, 'content', 'delete')):
        messages.error(request, 'You do not have permission to delete services.')
        return redirect('panel_services')
    
    get_object_or_404(Service, pk=pk).delete()
    messages.success(request, 'Service deleted.')
    return redirect('panel_services')


#   Why Choose Us  

@login_required(login_url='panel_login')
@permission_required('content', 'create')
def panel_why_add(request):
    if request.method == 'POST':
        WhyChooseUs.objects.create(
            title=request.POST['title'], description=request.POST['description'],
            icon=request.POST.get('icon', ''), order=request.POST.get('order', 0),
            is_active=request.POST.get('is_active') == 'on',
        )
        messages.success(request, 'Item added.')
        return redirect('panel_services')
    return render(request, 'panel/why_form.html', {'action': 'Add'})

@login_required(login_url='panel_login')
@permission_required('content', 'edit')
def panel_why_edit(request, pk):
    item = get_object_or_404(WhyChooseUs, pk=pk)
    if request.method == 'POST':
        item.title = request.POST['title']
        item.description = request.POST['description']
        item.icon = request.POST.get('icon', '')
        item.order = request.POST.get('order', 0)
        item.is_active = request.POST.get('is_active') == 'on'
        item.save()
        messages.success(request, 'Item updated.')
        return redirect('panel_services')
    return render(request, 'panel/why_form.html', {'action': 'Edit', 'item': item})

@login_required(login_url='panel_login')
@permission_required('content', 'delete')
def panel_why_delete(request, pk):
    get_object_or_404(WhyChooseUs, pk=pk).delete()
    messages.success(request, 'Item deleted.')
    return redirect('panel_services')


#   About Page  

@login_required(login_url='panel_login')
@permission_required('content', 'view')
def panel_about(request):
    can_create = request.user.is_superuser or check_permission(request.user, 'content', 'create')
    can_edit = request.user.is_superuser or check_permission(request.user, 'content', 'edit')
    can_delete = request.user.is_superuser or check_permission(request.user, 'content', 'delete')
    
    return render(request, 'panel/about.html', {
        'stats': Stat.objects.all(),
        'trusted_clients': TrustedClient.objects.all(),
        'testimonials': Testimonial.objects.all(),
        'team': TeamMember.objects.all(),
        'founders': Founder.objects.all(),
        'about_content': AboutContent.objects.all(),
        'can_create': can_create,
        'can_edit': can_edit,
        'can_delete': can_delete,
    })

# About Content Management
@login_required(login_url='panel_login')
@permission_required('content', 'view')
def panel_about_content_save(request, pk=None):
    about_content = get_object_or_404(AboutContent, pk=pk) if pk else None
    
    # Check create/edit permission
    if request.method == 'POST':
        required_perm = 'edit' if about_content else 'create'
        if not (request.user.is_superuser or check_permission(request.user, 'content', required_perm)):
            messages.error(request, f'You do not have permission to {required_perm} about content.')
            return redirect('panel_about')
        data = dict(
            mission_title=request.POST['mission_title'],
            mission_content=request.POST['mission_content'],
            vision_title=request.POST['vision_title'],
            vision_content=request.POST['vision_content'],
            quote_content=request.POST['quote_content'],
            quote_author=request.POST.get('quote_author', ''),
            order=request.POST.get('order', 0),
            is_active=request.POST.get('is_active') == 'on'
        )
        if about_content:
            for k, v in data.items(): setattr(about_content, k, v)
            about_content.save()
        else:
            about_content = AboutContent.objects.create(**data)
        messages.success(request, 'About content saved.')
        return redirect('panel_about')
    return render(request, 'panel/about_content_form.html', {'action': 'Edit' if about_content else 'Add', 'about_content': about_content})

@login_required(login_url='panel_login')
@permission_required('content', 'view')
def panel_about_content_delete(request, pk):
    # Check delete permission
    if not (request.user.is_superuser or check_permission(request.user, 'content', 'delete')):
        messages.error(request, 'You do not have permission to delete about content.')
        return redirect('panel_about')
    
    get_object_or_404(AboutContent, pk=pk).delete()
    messages.success(request, 'About content deleted.')
    return redirect('panel_about')

# Founder Management
@login_required(login_url='panel_login')
@permission_required('content', 'view')
def panel_founder_save(request, pk=None):
    founder = get_object_or_404(Founder, pk=pk) if pk else None
    
    # Check create/edit permission
    if request.method == 'POST':
        required_perm = 'edit' if founder else 'create'
        if not (request.user.is_superuser or check_permission(request.user, 'content', required_perm)):
            messages.error(request, f'You do not have permission to {required_perm} founder.')
            return redirect('panel_about')
        data = dict(name=request.POST['name'], role=request.POST['role'],
                    phone=request.POST.get('phone', ''), bio=request.POST.get('bio', ''),
                    order=request.POST.get('order', 0), is_active=request.POST.get('is_active') == 'on')
        if founder:
            for k, v in data.items(): setattr(founder, k, v)
            if 'photo' in request.FILES: founder.photo = request.FILES['photo']
            founder.save()
        else:
            founder = Founder.objects.create(**data)
            if 'photo' in request.FILES: founder.photo = request.FILES['photo']; founder.save()
        messages.success(request, 'Founder saved.')
        return redirect('panel_about')
    return render(request, 'panel/founder_form.html', {'action': 'Edit' if founder else 'Add', 'founder': founder})

@login_required(login_url='panel_login')
@permission_required('content', 'view')
def panel_founder_delete(request, pk):
    # Check delete permission
    if not (request.user.is_superuser or check_permission(request.user, 'content', 'delete')):
        messages.error(request, 'You do not have permission to delete founder.')
        return redirect('panel_about')
    
    get_object_or_404(Founder, pk=pk).delete()
    messages.success(request, 'Founder deleted.')
    return redirect('panel_about')

# Stats
@login_required(login_url='panel_login')
@permission_required('content', 'view')
def panel_stat_save(request, pk=None):
    stat = get_object_or_404(Stat, pk=pk) if pk else None
    
    # Check create/edit permission
    if request.method == 'POST':
        required_perm = 'edit' if stat else 'create'
        if not (request.user.is_superuser or check_permission(request.user, 'content', required_perm)):
            messages.error(request, f'You do not have permission to {required_perm} stats.')
            return redirect('panel_about')
        data = dict(value=request.POST['value'], label=request.POST['label'],
                    order=request.POST.get('order', 0), is_active=request.POST.get('is_active') == 'on')
        if stat:
            for k, v in data.items(): setattr(stat, k, v)
            stat.save()
        else:
            Stat.objects.create(**data)
        messages.success(request, 'Stat saved.')
        return redirect('panel_about')
    return render(request, 'panel/stat_form.html', {'action': 'Edit' if stat else 'Add', 'stat': stat})

@login_required(login_url='panel_login')
@permission_required('content', 'view')
def panel_stat_delete(request, pk):
    # Check delete permission
    if not (request.user.is_superuser or check_permission(request.user, 'content', 'delete')):
        messages.error(request, 'You do not have permission to delete stats.')
        return redirect('panel_about')
    
    get_object_or_404(Stat, pk=pk).delete()
    messages.success(request, 'Stat deleted.')
    return redirect('panel_about')

# Trusted Clients
@login_required(login_url='panel_login')
@permission_required('content', 'edit')
def panel_trusted_save(request, pk=None):
    obj = get_object_or_404(TrustedClient, pk=pk) if pk else None
    if request.method == 'POST':
        data = dict(name=request.POST['name'], icon=request.POST.get('icon', ''),
                    order=request.POST.get('order', 0), is_active=request.POST.get('is_active') == 'on')
        if obj:
            for k, v in data.items(): setattr(obj, k, v)
            if 'logo' in request.FILES: obj.logo = request.FILES['logo']
            obj.save()
        else:
            obj = TrustedClient.objects.create(**data)
            if 'logo' in request.FILES: obj.logo = request.FILES['logo']; obj.save()
        messages.success(request, 'Client saved.')
        return redirect('panel_about')
    return render(request, 'panel/trusted_form.html', {'action': 'Edit' if obj else 'Add', 'obj': obj})

@login_required(login_url='panel_login')
@permission_required('content', 'view')
def panel_trusted_delete(request, pk):
    # Check delete permission
    if not (request.user.is_superuser or check_permission(request.user, 'content', 'delete')):
        messages.error(request, 'You do not have permission to delete clients.')
        return redirect('panel_about')
    
    get_object_or_404(TrustedClient, pk=pk).delete()
    messages.success(request, 'Client deleted.')
    return redirect('panel_about')

# Testimonials
@login_required(login_url='panel_login')
@permission_required('content', 'edit')
def panel_testimonial_save(request, pk=None):
    obj = get_object_or_404(Testimonial, pk=pk) if pk else None
    if request.method == 'POST':
        data = dict(quote=request.POST['quote'], author_name=request.POST['author_name'],
                    author_role=request.POST.get('author_role', ''), initials=request.POST.get('initials', ''),
                    order=request.POST.get('order', 0), is_active=request.POST.get('is_active') == 'on')
        if obj:
            for k, v in data.items(): setattr(obj, k, v)
            if 'photo' in request.FILES: obj.photo = request.FILES['photo']
            obj.save()
        else:
            obj = Testimonial.objects.create(**data)
            if 'photo' in request.FILES: obj.photo = request.FILES['photo']; obj.save()
        messages.success(request, 'Testimonial saved.')
        return redirect('panel_about')
    return render(request, 'panel/testimonial_form.html', {'action': 'Edit' if obj else 'Add', 'obj': obj})

@login_required(login_url='panel_login')
@permission_required('content', 'view')
def panel_testimonial_delete(request, pk):
    # Check delete permission
    if not (request.user.is_superuser or check_permission(request.user, 'content', 'delete')):
        messages.error(request, 'You do not have permission to delete testimonials.')
        return redirect('panel_about')
    
    get_object_or_404(Testimonial, pk=pk).delete()
    messages.success(request, 'Testimonial deleted.')
    return redirect('panel_about')

# Team Members
@login_required(login_url='panel_login')
@permission_required('content', 'edit')
def panel_team_save(request, pk=None):
    obj = get_object_or_404(TeamMember, pk=pk) if pk else None
    if request.method == 'POST':
        data = dict(name=request.POST['name'], role=request.POST['role'],
                    phone=request.POST.get('phone', ''), bio=request.POST.get('bio', ''),
                    order=request.POST.get('order', 0), is_active=request.POST.get('is_active') == 'on')
        if obj:
            for k, v in data.items(): setattr(obj, k, v)
            if 'photo' in request.FILES: obj.photo = request.FILES['photo']
            obj.save()
        else:
            obj = TeamMember.objects.create(**data)
            if 'photo' in request.FILES: obj.photo = request.FILES['photo']; obj.save()
        messages.success(request, 'Team member saved.')
        return redirect('panel_about')
    return render(request, 'panel/team_form.html', {'action': 'Edit' if obj else 'Add', 'obj': obj})

@login_required(login_url='panel_login')
@permission_required('content', 'view')
def panel_team_delete(request, pk):
    # Check delete permission
    if not (request.user.is_superuser or check_permission(request.user, 'content', 'delete')):
        messages.error(request, 'You do not have permission to delete team members.')
        return redirect('panel_about')
    
    get_object_or_404(TeamMember, pk=pk).delete()
    messages.success(request, 'Team member deleted.')
    return redirect('panel_about')


#   Contact Inquiries  

@login_required(login_url='panel_login')
@permission_required('content', 'view')
def panel_inquiries(request):
    inquiries = ContactInquiry.objects.all()
    # mark as read when viewed
    inquiries.filter(is_read=False).update(is_read=True)
    
    can_delete = request.user.is_superuser or check_permission(request.user, 'content', 'delete')
    
    return render(request, 'panel/inquiries.html', {
        'inquiries': inquiries,
        'can_delete': can_delete,
    })

@login_required(login_url='panel_login')
@permission_required('content', 'view')
def panel_inquiry_delete(request, pk):
    # Check delete permission
    if not (request.user.is_superuser or check_permission(request.user, 'content', 'delete')):
        messages.error(request, 'You do not have permission to delete inquiries.')
        return redirect('panel_inquiries')
    
    get_object_or_404(ContactInquiry, pk=pk).delete()
    messages.success(request, 'Inquiry deleted.')
    return redirect('panel_inquiries')


#   Quote Requests  

from .models import QuoteRequest as QuoteRequestModel, QuotationRequest, QuotationRequestItem

@login_required(login_url='panel_login')
@permission_required('quotations', 'view')
def panel_quotes(request):
    quotations = QuotationRequest.objects.select_related('linked_customer__tier').prefetch_related('items__product__tier_prices__tier').all()
    tiers = CustomerTier.objects.filter(is_active=True)
    
    can_edit = request.user.is_superuser or check_permission(request.user, 'quotations', 'edit')
    can_delete = request.user.is_superuser or check_permission(request.user, 'quotations', 'delete')
    
    return render(request, 'panel/quotations.html', {
        'quotations': quotations,
        'tiers': tiers,
        'can_edit': can_edit,
        'can_delete': can_delete,
    })

@login_required(login_url='panel_login')
@permission_required('quotations', 'edit')
def panel_quote_update(request, pk):
    quotation = get_object_or_404(QuotationRequest, pk=pk)
    if request.method == 'POST':
        quotation.status = request.POST.get('status', quotation.status)
        quotation.admin_note = request.POST.get('admin_note', '')
        quotation.save()
        messages.success(request, 'Quotation updated.')
    return redirect('panel_quotes')

@login_required(login_url='panel_login')
@permission_required('quotations', 'edit')
def panel_quote_create_customer(request, pk):
    quotation = get_object_or_404(QuotationRequest, pk=pk)
    tiers = CustomerTier.objects.filter(is_active=True)
    existing_customer = Customer.objects.filter(email=quotation.user_email).first()
    if request.method == 'POST':
        tier_id = request.POST.get('tier') or None
        if existing_customer:
            if tier_id:
                existing_customer.tier_id = tier_id
                existing_customer.save()
            messages.success(request, 'Customer tier updated.')
            customer = existing_customer
        else:
            customer = Customer.objects.create(
                name=quotation.user_name,
                email=quotation.user_email,
                phone=quotation.phone,
                tier_id=tier_id,
            )
            messages.success(request, f'Customer "{customer.name}" created.')
        quotation.linked_customer = customer
        quotation.status = 'responded'
        quotation.save()
        return redirect('panel_quotes')
    selected_tier = existing_customer.tier_id if existing_customer else None
    return render(request, 'panel/quote_create_customer.html', {
        'quotation': quotation,
        'tiers': tiers,
        'existing_customer': existing_customer,
        'selected_tier': selected_tier,
    })

@login_required(login_url='panel_login')
@permission_required('quotations', 'delete')
def panel_quote_delete(request, pk):
    get_object_or_404(QuotationRequest, pk=pk).delete()
    messages.success(request, 'Quotation deleted.')
    return redirect('panel_quotes')


#   Orders  

@login_required(login_url='panel_login')
@permission_required('orders', 'view')
def panel_orders(request):
    """Renders the Orders Management page shell (no data   loaded via AJAX)."""
    can_edit   = request.user.is_superuser or check_permission(request.user, 'orders', 'edit')
    can_delete = request.user.is_superuser or check_permission(request.user, 'orders', 'delete')
    
    # Check if user can mark availability
    restricted_roles = ['customer', 'admin_logistics', 'admin_purchase']
    can_mark_available = request.user.is_superuser or (
        not request.user.role or request.user.role.name not in restricted_roles
    )
    
    return render(request, 'panel/orders.html', {
        'can_edit':   can_edit,
        'can_delete': can_delete,
        'can_mark_available': can_mark_available,
    })
 
@login_required(login_url='panel_login')
@permission_required('orders', 'view')
@require_GET
def panel_orders_data(request):
    """
    AJAX endpoint   returns paginated, filtered order rows as JSON.
 
    Query params
    ------------
    search            searches order_number, full_name, phone, email
    status            order status slug  ('' = all)
    payment_status    unpaid | partial | paid  ('' = all)
    payment_method    cod | online | pickup    ('' = all)
    date_from         YYYY-MM-DD
    date_to           YYYY-MM-DD
    page              int (default 1)
    per_page          int (default 20, max 100)
    """
    from django.core.paginator import Paginator
    from django.db.models import Q
    import datetime
 
    qs = (
        Order.objects
        .select_related('user', 'referred_agent')
        .prefetch_related('items')
        .order_by('-created_at')
    )

    # Restrict admin_purchase to only processing orders
    if request.user.role and request.user.role.name == 'admin_purchase':
        qs = qs.filter(status='processing')
 
    # -- filters ---------------------------------------------------------------
    search = request.GET.get('search', '').strip()
    if search:
        qs = qs.filter(
            Q(order_number__icontains=search)
            | Q(full_name__icontains=search)
            | Q(phone__icontains=search)
            | Q(email__icontains=search)
        )
 
    status = request.GET.get('status', '').strip()
    if status:
        qs = qs.filter(status=status)
 
    payment_status = request.GET.get('payment_status', '').strip()
    if payment_status:
        qs = qs.filter(payment_status=payment_status)
 
    payment_method = request.GET.get('payment_method', '').strip()
    if payment_method:
        qs = qs.filter(payment_method=payment_method)
 
    date_from = request.GET.get('date_from', '').strip()
    if date_from:
        try:
            qs = qs.filter(created_at__date__gte=datetime.date.fromisoformat(date_from))
        except ValueError:
            pass
 
    date_to = request.GET.get('date_to', '').strip()
    if date_to:
        try:
            qs = qs.filter(created_at__date__lte=datetime.date.fromisoformat(date_to))
        except ValueError:
            pass
 
    # -- pagination ------------------------------------------------------------
    try:
        per_page = min(int(request.GET.get('per_page', 20)), 100)
    except ValueError:
        per_page = 20
 
    paginator  = Paginator(qs, per_page)
    page_num   = request.GET.get('page', 1)
    page_obj   = paginator.get_page(page_num)
 
    can_edit = request.user.is_superuser or check_permission(request.user, 'orders', 'edit')
 
    # -- serialise rows --------------------------------------------------------
    rows = []
    for order in page_obj:
        agent = order.referred_agent
        rows.append({
            'pk':             order.pk,
            'order_number':   order.order_number,
            'full_name':      order.full_name,
            'email':          order.email or ' ',
            'phone':          order.phone or ' ',
            'items_count':    order.items.count(),
            'total':          str(order.total),
            'status':         order.status,
            'payment_status': order.payment_status,
            'payment_method': order.payment_method,
            'delivery_type':  order.delivery_type,
            'is_package':     order.is_package_order,
            'package_name':   order.package_name or '',
            'agent':          (agent.get_full_name() or agent.email) if agent else None,
            'created_at':     order.created_at.strftime('%d %b %Y, %H:%M'),
            'can_edit':       can_edit,
            'is_available':   order.is_avilable,
        })
 
    return JsonResponse({
        'rows':        rows,
        'total':       paginator.count,
        'page':        page_obj.number,
        'num_pages':   paginator.num_pages,
        'per_page':    per_page,
        'has_prev':    page_obj.has_previous(),
        'has_next':    page_obj.has_next(),
    })
 



@login_required(login_url='panel_login')
@permission_required('orders', 'edit')
def panel_order_detail(request, pk):
    from .models import ProductReview, OrderPayment
    order = get_object_or_404(Order.objects.select_related('user', 'referred_agent').prefetch_related('items__product', 'payments').order_by('-created_at'), pk=pk)
    
    # Get billing records for this order
    from django.db.models import Sum
    order_billings = Billing.objects.filter(
        items__product__in=[item.product for item in order.items.all() if item.product]
    ).distinct()
    
    # Calculate total paid from billings linked to this order
    total_paid_amount = 0
    related_bills = []
    
    # Better approach: Check billings created after this order with matching items
    for billing in Billing.objects.filter(created_at__gte=order.created_at).order_by('created_at'):
        # Check if billing items match order items
        billing_product_ids = set(billing.items.values_list('product_id', flat=True))
        order_product_ids = set(order.items.values_list('product_id', flat=True))
        
        # If there's significant overlap, consider it related
        if billing_product_ids & order_product_ids:  # Intersection
            related_bills.append(billing)
            total_paid_amount += float(billing.amount_paid)
    
    # Add payments from OrderPayment model
    order_payments = order.payments.all()
    for payment in order_payments:
        total_paid_amount += float(payment.amount)
    
    # Recalculate and update payment status based on actual total paid
    order_total = float(order.total)
    if order.payment_status != 'refunded':
        if total_paid_amount >= order_total:
            if order.payment_status != 'paid':
                order.payment_status = 'paid'
                order.save()
        elif total_paid_amount > 0:
            if order.payment_status != 'partial':
                order.payment_status = 'partial'
                order.save()
        else:
            if order.payment_status != 'unpaid':
                order.payment_status = 'unpaid'
                order.save()
    
    if request.method == 'POST':
        old_status = order.status
        new_status = request.POST.get('status', order.status)
        
        # Validate status transition using OrderStatusPermission
        if old_status != new_status:
            if not request.user.can_update_order_status(old_status, new_status):
                messages.error(request, f'You do not have permission to change order status from {old_status} to {new_status}.')
                return redirect('panel_order_detail', pk=pk)
        
        order.status = new_status
        order.save()
        if old_status != new_status:
            send_order_status_update_email(order, old_status, new_status)
        if old_status != 'delivered' and new_status == 'delivered':
            for item in order.items.all():
                if item.product:
                    StockEntry.objects.create(
                        product=item.product,
                        entry_type='sale',
                        quantity_change=-item.quantity,
                        note=f'Order {order.order_number} delivered',
                    )
        messages.success(request, 'Order status updated.')
        return redirect('panel_order_detail', pk=pk)
    
    # Get reviews for delivered orders
    items_with_reviews = []
    for item in order.items.all():
        review = None
        if item.product and order.status == 'delivered':
            review = ProductReview.objects.filter(product=item.product, user=order.user, order=order).first()
        items_with_reviews.append({'item': item, 'review': review})
    
    # Calculate tax breakdown
    tax_breakdown = {
        'has_tax': False,
        'total_tax': 0,
        'taxable_amount': 0,
        'avg_tax_rate': 0
    }
    
    total_tax = 0
    taxable_amount = 0
    tax_items_count = 0
    total_tax_rate = 0
    
    for item in order.items.all():
        if item.product and item.product.tax_included:
            item_total = float(item.subtotal)           
            tax_rate = 13.0  # Hardcoded 13% VAT
            item_tax = item_total * (tax_rate / (100 + tax_rate))
            item_taxable = item_total - item_tax
            
            total_tax += item_tax
            taxable_amount += item_taxable
            total_tax_rate += tax_rate
            tax_items_count += 1
    
    if tax_items_count > 0:
        tax_breakdown['has_tax'] = True
        tax_breakdown['total_tax'] = total_tax
        tax_breakdown['taxable_amount'] = taxable_amount
        tax_breakdown['avg_tax_rate'] = total_tax_rate / tax_items_count
    
    can_edit = request.user.is_superuser or check_permission(request.user, 'orders', 'edit')
    can_delete = request.user.is_superuser or check_permission(request.user, 'orders', 'delete')
    
    # Check if user can mark availability (only superadmin, admin_purchase, admin_sales)
    can_mark_available = request.user.is_superuser or (
        request.user.role and request.user.role.name in ['admin_purchase', 'admin_sales']
    )
    
    # Show availability toggle button based on role and status
    # admin_purchase can toggle in processing, others in pending
    show_availability_toggle = can_mark_available and (
        order.status == 'pending' or 
        (order.status == 'processing' and request.user.role and request.user.role.name == 'admin_purchase')
    )
    
    # Check if order status can be updated
    # Superadmin and admin_sales can bypass availability check for pending orders
    can_bypass_availability = request.user.is_superuser or (
        request.user.role and request.user.role.name == 'admin_sales'
    )
    can_update_status = can_edit and (order.is_avilable or order.status != 'pending' or can_bypass_availability)
    
    # Get allowed status transitions for current user
    allowed_statuses = []
    if can_update_status:
        if request.user.is_superuser:
            allowed_statuses = [status[0] for status in Order.STATUS_CHOICES]
        elif request.user.role:
            allowed_statuses = list(
                request.user.role.order_status_permissions
                .filter(from_status=order.status)
                .values_list('to_status', flat=True)
            )
            # Always include current status in the list
            if order.status not in allowed_statuses:
                allowed_statuses.append(order.status)
    
    return render(request, 'panel/order_detail.html', {
        'order': order,
        'items_with_reviews': items_with_reviews,
        'can_edit': can_edit,
        'can_delete': can_delete,
        'can_update_status': can_update_status,
        'allowed_statuses': allowed_statuses,
        'total_paid_amount': total_paid_amount,
        'remaining_amount': max(0, order_total - total_paid_amount),
        'related_bills': related_bills,
        'order_payments': order_payments,
        'tax_breakdown': tax_breakdown,
        'can_mark_available': can_mark_available,
        'show_availability_toggle': show_availability_toggle,
    })


@login_required(login_url='panel_login')
@permission_required('orders', 'view')
def panel_order_receipt(request, pk):
    order = get_object_or_404(
        Order.objects.select_related('user', 'referred_agent').prefetch_related('items__product').order_by('-created_at'),
        pk=pk
    )
    settings = SiteSettings.get()
    
    tax_breakdown = {'has_tax': False, 'total_tax': 0, 'taxable_amount': 0, 'avg_tax_rate': 0}
    total_tax = taxable_amount = tax_items_count = total_tax_rate = 0
    items_with_tax = []
    
    for item in order.items.all():
        item_data = {'item': item, 'unit_price_excl': 0, 'taxable_amount': 0, 'vat_amount': 0, 'total_amount': float(item.subtotal)}
        
        if item.product and item.product.tax_included:
            item_total = float(item.subtotal)
            tax_rate = 13.0
            item_tax = item_total * (tax_rate / (100 + tax_rate))
            item_taxable = item_total - item_tax
            
            item_data['unit_price_excl'] = float(item.unit_price) * (100 / 113)
            item_data['taxable_amount'] = item_taxable
            item_data['vat_amount'] = item_tax
            
            total_tax += item_tax
            taxable_amount += item_taxable
            total_tax_rate += tax_rate
            tax_items_count += 1
        else:
            item_data['unit_price_excl'] = float(item.unit_price)
            item_data['taxable_amount'] = float(item.subtotal)
        
        items_with_tax.append(item_data)
    
    if tax_items_count > 0:
        tax_breakdown.update({'has_tax': True, 'total_tax': total_tax, 'taxable_amount': taxable_amount, 'avg_tax_rate': total_tax_rate / tax_items_count})
    
    return render(request, 'panel/order_receipt.html', {'order': order, 'settings': settings, 'tax_breakdown': tax_breakdown, 'items_with_tax': items_with_tax})
@login_required(login_url='panel_login')
@permission_required('orders', 'delete')
def panel_order_delete(request, pk):
    order = get_object_or_404(Order, pk=pk)
    order_number = order.order_number
    order.delete()
    messages.success(request, f'Order {order_number} deleted.')
    return redirect('panel_orders')


@login_required(login_url='panel_login')
@permission_required('orders', 'edit')
def toggle_order_availability(request, pk):
    """Toggle order availability status - only for superadmin, admin_purchase, admin_sales."""
    
    # Check if user has permission to mark availability
    if not (request.user.is_superuser or (
        request.user.role and request.user.role.name in ['admin_purchase', 'admin_sales']
    )):
        messages.error(request, 'You do not have permission to mark order availability.')
        return redirect('panel_order_detail', pk=pk)
    
    order = get_object_or_404(Order, pk=pk)
    
    # Only allow toggling for pending orders
    if order.status != 'pending':
        messages.error(request, 'Availability can only be toggled for pending orders.')
        return redirect('panel_order_detail', pk=pk)
    
    order.is_avilable = not order.is_avilable
    order.save()
    
    status_text = "available" if order.is_avilable else "unavailable"
    messages.success(request, f'Order #{order.order_number} marked as {status_text}.')
    
    # Return JSON for AJAX or redirect for regular request
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'success': True, 'is_available': order.is_avilable})
    return redirect('panel_order_detail', pk=pk)


@login_required(login_url='panel_login')
@permission_required('orders', 'edit')
def mark_order_as_paid(request, pk):
    from .models import OrderPayment
    
    if not (request.user.is_superuser or (request.user.role and request.user.role.name in ['admin_finance', 'admin_sales'])):
        messages.error(request, 'You do not have permission to mark orders as paid.')
        return redirect('panel_order_detail', pk=pk)
    
    order = get_object_or_404(Order, pk=pk)
    
    # Calculate remaining amount
    from django.db.models import Sum
    total_paid = order.payments.aggregate(total=Sum('amount'))['total'] or 0
    remaining = float(order.total) - float(total_paid)
    
    if remaining > 0:
        # Create payment record for remaining amount
        OrderPayment.objects.create(
            order=order,
            amount=remaining,
            payment_method='cash',
            note='Marked as paid by admin',
            recorded_by=request.user
        )
        messages.success(request, f'Order #{order.order_number} marked as paid with payment of Rs. {remaining:.2f} recorded.')
    else:
        messages.info(request, f'Order #{order.order_number} is already fully paid.')
    
    return redirect('panel_order_detail', pk=pk)


@login_required(login_url='panel_login')
@permission_required('orders', 'edit')
def record_order_payment(request, pk):
    import json as _json

@login_required(login_url='panel_login')
@permission_required('orders', 'edit')


@login_required(login_url='panel_login')
@permission_required('orders', 'edit')
def mark_order_as_refund(request, pk):
    from .models import OrderPayment
    
    if not (request.user.is_superuser or (request.user.role and request.user.role.name in ['admin_finance', 'admin_sales'])):
        messages.error(request, 'You do not have permission to mark orders as refund.')
        return redirect('panel_order_detail', pk=pk)
    
    order = get_object_or_404(Order, pk=pk)
    deleted_count = order.payments.count()
    order.payments.all().delete()
    Order.objects.filter(pk=pk).update(payment_status='refunded')   
    messages.success(request, f'Order #{order.order_number} marked as refunded. {deleted_count} payment record(s) removed.')
    return redirect('panel_order_detail', pk=pk)

@login_required(login_url='panel_login')
@permission_required('orders', 'edit')
def toggle_order_item_availability(request, pk):
    from .models import OrderItem
    
    # Check if user has permission to mark availability
    if not (request.user.is_superuser or (
        request.user.role and request.user.role.name in ['admin_purchase', 'admin_sales']
    )):
        messages.error(request, 'You do not have permission to mark item availability.')
        return redirect('panel_orders')
    
    item = get_object_or_404(OrderItem, pk=pk)
    order = item.order
    
    # Allow toggling for pending and processing orders
    # admin_purchase can toggle in processing status
    if order.status not in ['pending', 'processing']:
        messages.error(request, 'Item availability can only be toggled for pending or processing orders.')
        return redirect('panel_order_detail', pk=order.pk)
    
    # Additional check: admin_purchase can only toggle in processing status
    if request.user.role and request.user.role.name == 'admin_purchase' and order.status != 'processing':
        messages.error(request, 'You can only toggle item availability for processing orders.')
        return redirect('panel_order_detail', pk=order.pk)
    
    item.is_available = not item.is_available
    item.save()
    
    # Auto-mark order as available if all items are available
    all_items_available = order.items.filter(is_available=False).count() == 0
    if all_items_available:
        order.is_avilable = True
        order.save()
    
    return redirect('panel_order_detail', pk=order.pk)


@login_required(login_url='panel_login')
@permission_required('orders', 'edit')
def record_order_payment(request, pk):
    import json as _json
    order = get_object_or_404(Order, pk=pk)
    
    try:
        data = _json.loads(request.body)
    except:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    
    amount = float(data.get('amount', 0))
    payment_method = data.get('payment_method', 'cash')
    note = data.get('note', '').strip()
    
    if amount <= 0:
        return JsonResponse({'error': 'Invalid amount'}, status=400)
    
    # Create payment record
    payment = OrderPayment.objects.create(
        order=order,
        amount=amount,
        payment_method=payment_method,
        note=note,
        recorded_by=request.user
    )
    
    # Calculate total paid from OrderPayment records
    from django.db.models import Sum
    total_paid_from_payments = order.payments.aggregate(total=Sum('amount'))['total'] or 0
    
    # Also check billings
    total_paid_from_billings = 0
    for billing in Billing.objects.filter(created_at__gte=order.created_at).order_by('created_at'):
        billing_product_ids = set(billing.items.values_list('product_id', flat=True))
        order_product_ids = set(order.items.values_list('product_id', flat=True))
        if billing_product_ids & order_product_ids:
            total_paid_from_billings += float(billing.amount_paid)
    
    total_paid = float(total_paid_from_payments) + float(total_paid_from_billings)
    
    # Update order payment status
    order_total = float(order.total)
    if total_paid >= order_total:
        order.payment_status = 'paid'
    elif total_paid > 0:
        order.payment_status = 'partial'
    else:
        order.payment_status = 'unpaid'
    order.save()
    
    return JsonResponse({
        'success': True,
        'payment_id': payment.pk,
        'total_paid': float(total_paid),
        'payment_status': order.payment_status
    })


#   User Management  

@login_required(login_url='panel_login')
@permission_required('users', 'view')
def panel_users(request):
    from django.core.paginator import Paginator
    search = request.GET.get('search', '').strip()
    role_id = request.GET.get('role', '').strip()
    user_type = request.GET.get('user_type', '').strip()
    page = request.GET.get('page', 1)
    
    # Show all users (staff and non-staff) to see agents and customers
    qs = CustomerUser.objects.select_related('role').all()
    
    if search:
        qs = qs.filter(username__icontains=search) | qs.filter(email__icontains=search)
    
    if role_id:
        try:
            qs = qs.filter(role_id=int(role_id))
        except:
            pass
    
    if user_type:
        qs = qs.filter(user_type=user_type)
    
    qs = qs.order_by('-date_joined')
    
    paginator = Paginator(qs, 20)
    users = paginator.get_page(page)
    roles = Role.objects.all()
    
    can_create = request.user.is_superuser or check_permission(request.user, 'users', 'create')
    can_edit = request.user.is_superuser or check_permission(request.user, 'users', 'edit')
    can_delete = request.user.is_superuser or check_permission(request.user, 'users', 'delete')
    
    return render(request, 'panel/users.html', {
        'users': users,
        'roles': list(roles),
        'search': search,
        'role_id': role_id,
        'user_type': user_type,
        'total_count': paginator.count,
        'can_create': can_create,
        'can_edit': can_edit,
        'can_delete': can_delete,
    })

@login_required(login_url='panel_login')
@permission_required('users', 'create')
def panel_user_add(request):
    roles = Role.objects.all()
    
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        email = request.POST.get('email', '').strip()
        password = request.POST.get('password', '')
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        phone = request.POST.get('phone', '').strip()
        is_staff = request.POST.get('is_staff') == 'on'
        is_active = request.POST.get('is_active') == 'on'
        user_type = request.POST.get('user_type', 'customer')
        customer_subtype = request.POST.get('customer_subtype', 'retailer')

        if not username or not email or not password:
            messages.error(request, 'Username, email, and password are required.')
            return render(request, 'panel/user_form.html', {
                'action': 'Add', 
                'roles': roles, 
                'user': None
            })
        
        if CustomerUser.objects.filter(username=username).exists():
            messages.error(request, 'Username already exists.')
            return render(request, 'panel/user_form.html', {
                'action': 'Add', 
                'roles': roles, 
                'user': None
            })
        
        if CustomerUser.objects.filter(email=email).exists():
            messages.error(request, 'Email already exists.')
            return render(request, 'panel/user_form.html', {
                'action': 'Add', 
                'roles': roles, 
                'user': None
            })

        # Get role from form or auto-assign based on user type
        role_id = request.POST.get('role')
        if role_id:
            role = Role.objects.filter(pk=role_id).first()
        else:
            # Auto-assign role based on user type
            if user_type == 'staff':
                role = Role.objects.filter(name__iexact="staff").first()
            else:
                role = Role.objects.filter(name__iexact="customer").first()
        
        if not role:
            role = Role.objects.first()
            if not role:
                messages.error(request, "No roles found in the system. Please create roles first.")
                return render(request, 'panel/user_form.html', {
                    'action': 'Add', 
                    'roles': roles, 
                    'user': None
                })

        # Create user
        user = CustomerUser.objects.create_user(
            username=username,
            email=email,
            password=password,
            first_name=first_name,
            last_name=last_name,
            phone=phone,
            role=role,
            is_staff=is_staff,
            is_active=is_active,
            is_superuser=False,
            user_type=user_type,
            customer_subtype=customer_subtype if user_type == 'customer' else None,
        )
        
        messages.success(request, f'User "{username}" created successfully.')
        return redirect('panel_users')
    
    return render(request, 'panel/user_form.html', {
        'action': 'Add', 
        'roles': roles, 
        'user': None
    })


@login_required(login_url='panel_login')
@permission_required('users', 'edit')
def panel_user_edit(request, pk):
    user = get_object_or_404(CustomerUser, pk=pk)
    roles = Role.objects.all()
    
    if request.method == 'POST':
        user.email = request.POST.get('email', user.email).strip()
        user.first_name = request.POST.get('first_name', '').strip()
        user.last_name = request.POST.get('last_name', '').strip()
        user.phone = request.POST.get('phone', '').strip()
        user.is_staff = request.POST.get('is_staff') == 'on'
        user.is_active = request.POST.get('is_active') == 'on'
        user.is_superuser = False
        user.user_type = request.POST.get('user_type', user.user_type)
        
        # Update customer subtype only if user is customer
        if user.user_type == 'customer':
            user.customer_subtype = request.POST.get('customer_subtype', 'retailer')
        else:
            user.customer_subtype = None

        # Allow changing role manually in edit mode
        role_id = request.POST.get('role')
        if role_id:
            user.role_id = role_id
        elif user.user_type == 'customer':
            # Auto-assign customer role if no role selected
            customer_role = Role.objects.filter(name__iexact="customer").first()
            if customer_role:
                user.role = customer_role

        password = request.POST.get('password', '').strip()
        if password:
            user.set_password(password)
        
        user.save()
        messages.success(request, f'User "{user.username}" updated successfully.')
        return redirect('panel_users')
    
    return render(request, 'panel/user_form.html', {
        'action': 'Edit', 
        'user': user, 
        'roles': roles
    })

@login_required(login_url='panel_login')
@permission_required('users', 'delete')
def panel_user_delete(request, pk):
    user = get_object_or_404(CustomerUser, pk=pk)
    username = user.username
    user.delete()
    messages.success(request, f'User "{username}" deleted.')
    return redirect('panel_users')


@login_required(login_url='panel_login')
@permission_required('users', 'view')
def panel_roles(request):
    can_edit = request.user.is_superuser or check_permission(request.user, 'users', 'edit')
    
    roles = Role.objects.prefetch_related('permissions').all()
    return render(request, 'panel/roles.html', {
        'roles': roles,
        'can_edit': can_edit,
    })

@login_required(login_url='panel_login')
@permission_required('users', 'edit')
def panel_role_edit(request, pk):
    role = get_object_or_404(Role, pk=pk)
    permissions = role.permissions.all()
    all_modules = Permission.MODULE_CHOICES
    all_actions = Permission.ACTION_CHOICES
    
    if request.method == 'POST':
        role.description = request.POST.get('description', '')
        role.save()
        
        # Update permissions
        selected_perms = request.POST.getlist('permissions')
        role.permissions.all().delete()
        
        for perm_str in selected_perms:
            try:
                module, action = perm_str.split('|')
                Permission.objects.create(role=role, module=module, action=action)
            except:
                pass
        
        messages.success(request, f'Role "{role.get_name_display()}" updated.')
        return redirect('panel_roles')
    
    current_perms = set(f"{p.module}|{p.action}" for p in permissions)
    
    return render(request, 'panel/role_edit.html', {
        'role': role,
        'all_modules': all_modules,
        'all_actions': all_actions,
        'current_perms': current_perms,
    })


from django.http import JsonResponse


# -- POS Billing ---------------------------------------------------------------

from .models import Billing, BillingItem

@login_required(login_url='panel_login')
@permission_required('orders', 'create')
def panel_billing(request):
    import json as _json
    from django.db.models import Sum, Count, Q
    from django.core.paginator import Paginator

    # Check if coming from order detail page
    order_id = request.GET.get('order_id')
    order_data = None
    if order_id:
        order = Order.objects.filter(pk=order_id).select_related('user').prefetch_related('items__product__images').first()
        if order:
            order_data = {
                'order_id': order.pk,
                'order_number': order.order_number,
                'customer_name': order.full_name,
                'customer_phone': order.phone,
                'customer_email': order.user.email if order.user else order.email,
                'is_package_order': order.is_package_order,
                'package_name': order.package_name if order.is_package_order else '',
                'total': float(order.total),
                'items': [{
                    'product_id': item.product_id,
                    'product_name': item.product_name,
                    'product_sku': item.product_sku,
                    'quantity': item.quantity,
                    'unit_price': float(item.unit_price),
                    'image': item.product.primary_image.image.url if item.product and item.product.primary_image else '',
                } for item in order.items.all()]
            }

    # -- POST: create a bill --------------------------------------------------
    if request.method == 'POST':
        try:
            data = _json.loads(request.body)
        except Exception:
            return JsonResponse({'ok': False, 'error': 'Invalid JSON'})

        items_data = data.get('items', [])
        if not items_data:
            return JsonResponse({'ok': False, 'error': 'No items'})

        sale_type        = data.get('sale_type', 'counter')
        customer_id      = data.get('customer_id') or None
        walk_in_name     = data.get('walk_in_name', '').strip()
        walk_in_phone    = data.get('walk_in_phone', '').strip()
        agent_id         = data.get('agent_id') or None
        overall_discount = abs(float(data.get('overall_discount', 0)))
        payment_method   = data.get('payment_method', 'cash')
        amount_paid      = abs(float(data.get('amount_paid', 0)))
        split_payments   = data.get('split_payments')
        note             = data.get('note', '').strip()
        linked_order_id  = data.get('order_id') or None

        # Calculate payment amounts based on method
        cash_amount = 0
        card_amount = 0
        online_amount = 0
        
        if payment_method == 'split' and split_payments:
            for sp in split_payments:
                amt = abs(float(sp.get('amount', 0)))
                method = sp.get('method', 'cash')
                if method == 'cash':
                    cash_amount += amt
                elif method == 'card':
                    card_amount += amt
                elif method == 'upi':
                    online_amount += amt
        else:
            if payment_method == 'cash':
                cash_amount = amount_paid
            elif payment_method == 'card':
                card_amount = amount_paid
            elif payment_method == 'upi':
                online_amount = amount_paid

        subtotal = 0
        item_discount_total = 0
        bill_items = []

        for it in items_data:
            product = Product.objects.filter(pk=it.get('product_id'), is_active=True).first()
            if not product:
                return JsonResponse({'ok': False, 'error': f"Product {it.get('product_id')} not found"})
            qty      = max(1, int(it.get('qty', 1)))
            price    = float(it.get('unit_price', float(product.mrp)))
            discount = abs(float(it.get('discount', 0)))
            line_sub = price * qty
            subtotal += line_sub
            item_discount_total += discount
            bill_items.append({
                'product': product, 'qty': qty,
                'price': price, 'discount': discount,
            })

        total       = max(0, subtotal - item_discount_total - overall_discount)
        amount_paid = cash_amount + card_amount + online_amount
        if amount_paid >= total:
            pay_status = 'paid'
        elif amount_paid > 0:
            pay_status = 'partial'
        else:
            pay_status = 'unpaid'

        bill = Billing.objects.create(
            sale_type=sale_type,
            customer_id=customer_id,
            walk_in_name=walk_in_name,
            walk_in_phone=walk_in_phone,
            agent_id=agent_id,
            billed_by=request.user,
            subtotal=subtotal,
            item_discount=item_discount_total,
            overall_discount=overall_discount,
            total=total,
            cash_amount=cash_amount,
            card_amount=card_amount,
            online_amount=online_amount,
            amount_paid=amount_paid,
            payment_status=pay_status,
            note=note,
        )

        for it in bill_items:
            BillingItem.objects.create(
                billing=bill,
                product=it['product'],
                product_name=it['product'].name,
                product_sku=it['product'].sku,
                unit_price=it['price'],
                quantity=it['qty'],
                discount=it['discount'],
            )
            # deduct stock
            StockEntry.objects.create(
                product=it['product'],
                entry_type='sale',
                quantity_change=-it['qty'],
                unit_price=it['price'],
                note=f'POS Bill #{bill.bill_number}',
            )

        # Update order payment status if linked
        if linked_order_id:
            order = Order.objects.filter(pk=linked_order_id).first()
            if order:
                # Update payment status based on amount paid vs total
                if amount_paid >= order.total:
                    order.payment_status = 'paid'
                elif amount_paid > 0:
                    order.payment_status = 'partial'
                else:
                    order.payment_status = 'unpaid'
                order.save()

        return JsonResponse({'ok': True, 'bill_number': bill.bill_number, 'bill_id': bill.pk})

    # -- GET: render POS page -------------------------------------------------
    # Products JSON for POS grid
    products_qs = Product.objects.filter(is_active=True).select_related('category').prefetch_related('images', 'tier_prices__tier')
    products_data = []
    for p in products_qs:
        img = p.primary_image
        tier_prices = {tp.tier_id: float(tp.price) for tp in p.tier_prices.all()}
        products_data.append({
            'id': p.pk, 'name': p.name, 'sku': p.sku,
            'mrp': float(p.mrp),
            'stock': p.stock_quantity,
            'category': p.category.name if p.category else '',
            'category_id': p.category_id or 0,
            'image': img.image.url if img else '',
            'tier_prices': tier_prices,
        })

    customers_qs = Customer.objects.filter(is_active=True).select_related('tier')
    customers_data = [{
        'id': c.pk, 'name': c.name, 'phone': c.phone, 'email': c.email,
        'tier_id': c.tier_id, 'tier_name': c.tier.name if c.tier else '',
    } for c in customers_qs]

    agents_data = [{
        'id': a.pk,
        'name': (a.get_full_name() or a.username),
        'referral_code': a.referral_code or '',
    } for a in CustomerUser.objects.filter(user_type='agent', is_active=True)]

    categories = list(Category.objects.filter(is_active=True).values('id', 'name').order_by('name'))

    # -- Bills table (paginated, filtered) ------------------------------------
    bills_qs = Billing.objects.select_related('customer', 'agent', 'billed_by').order_by('-created_at')

    f_search      = request.GET.get('search', '').strip()
    f_sale_type   = request.GET.get('sale_type', '').strip()
    f_pay_status  = request.GET.get('pay_status', '').strip()
    f_agent       = request.GET.get('agent', '').strip()
    f_date_from   = request.GET.get('date_from', '').strip()
    f_date_to     = request.GET.get('date_to', '').strip()
    page          = request.GET.get('page', 1)
    limit         = int(request.GET.get('limit', 20))

    if f_search:
        bills_qs = bills_qs.filter(
            Q(bill_number__icontains=f_search) |
            Q(walk_in_name__icontains=f_search) |
            Q(customer__name__icontains=f_search)
        )
    if f_sale_type:
        bills_qs = bills_qs.filter(sale_type=f_sale_type)
    if f_pay_status:
        bills_qs = bills_qs.filter(payment_status=f_pay_status)
    if f_agent:
        bills_qs = bills_qs.filter(agent_id=f_agent)
    if f_date_from:
        bills_qs = bills_qs.filter(created_at__date__gte=f_date_from)
    if f_date_to:
        bills_qs = bills_qs.filter(created_at__date__lte=f_date_to)

    paginator   = Paginator(bills_qs, limit)
    bills_page  = paginator.get_page(page)

    # -- Stats (unfiltered totals) ---------------------------------------------
    from decimal import Decimal
    stats = Billing.objects.aggregate(
        total_bills=Count('id'),
        total_revenue=Sum('total'),
        total_paid=Sum('amount_paid'),
    )
    
    # Calculate actual pending amount considering OrderPayments
    total_order_payments = OrderPayment.objects.aggregate(total=Sum('amount'))['total'] or Decimal('0')
    
    # Pending amount = Total Revenue - (Billing Payments + Order Payments)
    total_billed_paid = stats['total_paid'] or Decimal('0')
    total_revenue = stats['total_revenue'] or Decimal('0')
    actual_pending = total_revenue - total_billed_paid - total_order_payments

    # Check if user can create sales
    can_create_sale = request.user.is_superuser or (
        request.user.role and request.user.role.name in ['admin_finance', 'admin_sales']
    )
    
    return render(request, 'panel/billing.html', {
        'products_json':  _json.dumps(products_data),
        'customers_json': _json.dumps(customers_data),
        'agents_json':    _json.dumps(agents_data),
        'categories_json': _json.dumps(categories),
        'bills':          bills_page,
        'paginator':      paginator,
        'stats':          stats,
        'actual_pending': actual_pending,
        'f_search':       f_search,
        'f_sale_type':    f_sale_type,
        'f_pay_status':   f_pay_status,
        'f_agent':        f_agent,
        'f_date_from':    f_date_from,
        'f_date_to':      f_date_to,
        'limit':          limit,
        'order_data':     _json.dumps(order_data) if order_data else None,
        'can_create_sale': can_create_sale,
    })


@login_required(login_url='panel_login')
@permission_required('orders', 'view')
def panel_billing_view(request, pk):
    bill = get_object_or_404(
        Billing.objects.select_related('customer', 'agent', 'billed_by').prefetch_related('items__product'),
        pk=pk
    )
    settings = SiteSettings.get()
    return render(request, 'panel/billing_invoice.html', {
        'bill': bill,
        'settings': settings,
    })


@login_required(login_url='panel_login')
@permission_required('orders', 'view')
def panel_billing_detail(request, pk):
    bill = get_object_or_404(
        Billing.objects.select_related('customer', 'agent', 'billed_by').prefetch_related('items__product'),
        pk=pk
    )
    return JsonResponse({
        'bill_number': bill.bill_number,
        'sale_type': bill.get_sale_type_display(),
        'customer': bill.customer.name if bill.customer else bill.walk_in_name or 'Walk-in',
        'phone': bill.customer.phone if bill.customer else bill.walk_in_phone,
        'agent': (bill.agent.get_full_name() or bill.agent.username) if bill.agent else '',
        'billed_by': (bill.billed_by.get_full_name() or bill.billed_by.username) if bill.billed_by else '',
        'subtotal': float(bill.subtotal),
        'item_discount': float(bill.item_discount),
        'overall_discount': float(bill.overall_discount),
        'total': float(bill.total),
        'cash_amount': float(bill.cash_amount),
        'card_amount': float(bill.card_amount),
        'online_amount': float(bill.online_amount),
        'amount_paid': float(bill.amount_paid),
        'balance_due': float(bill.balance_due),
        'payment_status': bill.payment_status,
        'note': bill.note,
        'created_at': bill.created_at.strftime('%d %b %Y, %H:%M'),
        'items': [{
            'name': i.product_name, 'sku': i.product_sku,
            'qty': i.quantity, 'unit_price': float(i.unit_price),
            'discount': float(i.discount), 'subtotal': float(i.subtotal),
        } for i in bill.items.all()],
    })


@login_required(login_url='panel_login')
@permission_required('orders', 'view')
def api_billing_products(request):
    """Search products for POS   returns JSON."""
    q = request.GET.get('q', '').strip()
    customer_id = request.GET.get('customer_id')
    tier_id = None
    if customer_id:
        c = Customer.objects.filter(pk=customer_id).select_related('tier').first()
        if c and c.tier_id:
            tier_id = c.tier_id

    qs = Product.objects.filter(is_active=True).prefetch_related('images', 'tier_prices__tier')
    if q:
        from django.db.models import Q as DQ
        qs = qs.filter(DQ(name__icontains=q) | DQ(sku__icontains=q))
    results = []
    for p in qs[:30]:
        img = p.primary_image
        price = float(p.mrp)
        if tier_id:
            tp = p.tier_prices.filter(tier_id=tier_id).first()
            if tp:
                price = float(tp.price)
        results.append({
            'id': p.pk, 'name': p.name, 'sku': p.sku,
            'price': price, 'mrp': float(p.mrp),
            'stock': p.stock_quantity,
            'image': img.image.url if img else '',
        })
    return JsonResponse({'results': results})


@login_required(login_url='panel_login')
@permission_required('orders', 'view')
def api_user_get(request, pk):
    user = get_object_or_404(CustomerUser, pk=pk)
    return JsonResponse({
        'id': user.id,
        'username': user.username,
        'email': user.email,
        'first_name': user.first_name,
        'last_name': user.last_name,
        'phone': user.phone,
        'role_id': user.role_id,
        'user_type': user.user_type,
        'customer_subtype': user.customer_subtype,
        'is_active': user.is_active,
        'is_staff': user.is_staff,
    })


@login_required(login_url='panel_login')
@permission_required('orders', 'view')
def api_customers(request):
    customers = Customer.objects.filter(is_active=True).select_related('tier')
    return JsonResponse([{
        'id': c.pk, 'name': c.name, 'phone': c.phone, 'email': c.email,
        'tier_id': c.tier_id, 'tier_name': c.tier.name if c.tier else '',
    } for c in customers], safe=False)


@login_required(login_url='panel_login')
@permission_required('customers', 'create')
def api_customer_create(request):
    import json as _json
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    
    try:
        data = _json.loads(request.body)
    except:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    
    name = data.get('name', '').strip()
    phone = data.get('phone', '').strip()
    email = data.get('email', '').strip()
    pan_number = data.get('pan_number', '').strip().upper()
    
    if not name:
        return JsonResponse({'error': 'Name is required'}, status=400)
    
    # Check if customer already exists
    if email and Customer.objects.filter(email=email).exists():
        return JsonResponse({'error': 'Customer with this email already exists'}, status=400)
    
    customer = Customer.objects.create(
        name=name,
        phone=phone,
        email=email,
        pan_number=pan_number,
        is_active=True
    )
    
    return JsonResponse({
        'id': customer.pk,
        'name': customer.name,
        'phone': customer.phone,
        'email': customer.email,
        'pan_number': customer.pan_number
    })


@login_required(login_url='panel_login')
@permission_required('orders', 'view')
def api_agents(request):
    agents = CustomerUser.objects.filter(user_type='agent', is_active=True)
    return JsonResponse([{
        'id': a.pk, 'username': a.username,
        'name': a.get_full_name() or a.username,
    } for a in agents], safe=False)


@login_required(login_url='panel_login')
@permission_required('orders', 'view')
def api_billing_list(request):
    from django.db.models import Count, F, Value, Sum
    from django.db.models.functions import Coalesce
    from decimal import Decimal
    limit = int(request.GET.get('limit', 10))
    offset = int(request.GET.get('offset', 0))
    start_date = request.GET.get('start_date', '').strip()
    end_date = request.GET.get('end_date', '').strip()
    sale_type = request.GET.get('sale_type', '').strip()
    payment_status = request.GET.get('payment_status', '').strip()
    agent_id = request.GET.get('agent', '').strip()
    
    qs = Billing.objects.select_related('customer', 'agent').prefetch_related('items')
    
    # Apply filters
    if start_date:
        qs = qs.filter(created_at__date__gte=start_date)
    if end_date:
        qs = qs.filter(created_at__date__lte=end_date)
    if sale_type:
        qs = qs.filter(sale_type=sale_type)
    if payment_status:
        qs = qs.filter(payment_status=payment_status)
    if agent_id:
        qs = qs.filter(agent_id=agent_id)
    
    qs = qs.order_by('-created_at')
    count = qs.count()
    bills = qs[offset:offset+limit]
    
    # Calculate stats based on filtered queryset
    from decimal import Decimal
    from .models import OrderPayment
    filtered_bills = qs.all()
    total_bills = filtered_bills.count()
    total_revenue = sum(b.total for b in filtered_bills) or Decimal('0')
    
    # Calculate pending considering OrderPayments
    total_order_payments = OrderPayment.objects.aggregate(total=Sum('amount'))['total'] or Decimal('0')
    total_billed_paid = sum(b.amount_paid for b in filtered_bills) or Decimal('0')
    pending_amount = total_revenue - total_billed_paid - total_order_payments
    
    total_discount = sum(b.overall_discount + b.item_discount for b in filtered_bills) or Decimal('0')
    
    return JsonResponse({
        'count': count,
        'results': [{
            'id': b.pk,
            'bill_number': b.bill_number,
            'created_at': b.created_at.isoformat(),
            'customer_name': b.customer.name if b.customer else b.walk_in_name or 'Walk-in',
            'sale_type': b.sale_type,
            'items_count': b.items.count(),
            'total': float(b.total),
            'amount_paid': float(b.amount_paid),
            'payment_status': b.payment_status,
        } for b in bills],
        'statistics': {
            'total_bills': total_bills,
            'total_revenue': float(total_revenue),
            'pending_amount': float(pending_amount),
            'total_discount': float(total_discount),
        }
    })


@login_required(login_url='panel_login')
@permission_required('orders', 'create')
def api_billing_create(request):
    import json as _json
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    
    try:
        data = _json.loads(request.body)
    except:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    
    items = data.get('items', [])
    if not items:
        return JsonResponse({'error': 'No items'}, status=400)
    
    customer_id = data.get('customer_id')
    walk_in_name = data.get('walk_in_name', '').strip()
    walk_in_phone = data.get('walk_in_phone', '').strip()
    sale_type = data.get('sale_type', 'counter')
    agent_id = data.get('agent_id')
    payment_method = data.get('payment_method', 'cash')
    payment_status = data.get('payment_status', 'paid')
    amount_paid = float(data.get('amount_paid', 0))
    overall_discount = float(data.get('overall_discount', 0))
    split_payments = data.get('split_payments')
    order_id = data.get('order_id')

    # Order creation fields
    order_full_name = data.get('order_full_name', '').strip() or walk_in_name
    order_phone = data.get('order_phone', '').strip() or walk_in_phone
    order_email = data.get('order_email', '').strip()
    order_delivery_type = data.get('order_delivery_type', 'pickup')
    order_address = data.get('order_address', '').strip()
    order_city = data.get('order_city', '').strip()
    billing_org_name = data.get('billing_org_name', '').strip()
    billing_person_name = data.get('billing_person_name', '').strip()
    billing_contact = data.get('billing_contact', '').strip()
    billing_address = data.get('billing_address', '').strip()
    billing_city = data.get('billing_city', '').strip()
    pan_number = data.get('pan_number', '').strip()
    
    subtotal = 0
    item_discount_total = 0
    bill_items = []
    
    for item in items:
        product = Product.objects.filter(pk=item['product_id']).first()
        if not product:
            return JsonResponse({'error': f"Product {item['product_id']} not found"}, status=400)
        
        qty = int(item['quantity'])
        price = float(item['price'])
        discount_pct = float(item.get('discount', 0))
        
        line_total = price * qty
        line_discount = line_total * (discount_pct / 100)
        
        subtotal += line_total
        item_discount_total += line_discount
        
        bill_items.append({
            'product': product,
            'qty': qty,
            'price': price,
            'discount': line_discount,
        })
    
    overall_discount_amount = subtotal * (overall_discount / 100)
    total = subtotal - item_discount_total - overall_discount_amount
    
    cash_amount = 0
    card_amount = 0
    online_amount = 0
    
    if payment_method == 'split' and split_payments:
        for sp in split_payments:
            amt = float(sp['amount'])
            if sp['method'] == 'cash':
                cash_amount += amt
            elif sp['method'] == 'card':
                card_amount += amt
            elif sp['method'] == 'upi':
                online_amount += amt
    else:
        if payment_method == 'cash':
            cash_amount = amount_paid
        elif payment_method == 'card':
            card_amount = amount_paid
        elif payment_method == 'upi':
            online_amount = amount_paid
    
    total_paid = cash_amount + card_amount + online_amount
    
    bill = Billing.objects.create(
        sale_type=sale_type,
        customer_id=customer_id,
        walk_in_name=walk_in_name,
        walk_in_phone=walk_in_phone,
        agent_id=agent_id,
        billed_by=request.user,
        subtotal=subtotal,
        item_discount=item_discount_total,
        overall_discount=overall_discount_amount,
        total=total,
        cash_amount=cash_amount,
        card_amount=card_amount,
        online_amount=online_amount,
        amount_paid=total_paid,
        payment_status=payment_status,
    )
    
    for item in bill_items:
        BillingItem.objects.create(
            billing=bill,
            product=item['product'],
            product_name=item['product'].name,
            product_sku=item['product'].sku,
            unit_price=item['price'],
            quantity=item['qty'],
            discount=item['discount'],
        )
        StockEntry.objects.create(
            product=item['product'],
            entry_type='sale',
            quantity_change=-item['qty'],
            unit_price=item['price'],
            note=f'Bill #{bill.bill_number}',
        )

    # Determine payment_method for Order
    order_payment_method = 'cod'
    if payment_method in ('card', 'upi', 'online'):
        order_payment_method = 'online'
    elif order_delivery_type == 'pickup':
        order_payment_method = 'pickup'

    # Determine order payment_status
    if total_paid >= total:
        order_pay_status = 'paid'
    elif total_paid > 0:
        order_pay_status = 'partial'
    else:
        order_pay_status = 'unpaid'

    # Resolve agent for order
    referred_agent = None
    if agent_id:
        referred_agent = CustomerUser.objects.filter(pk=agent_id, user_type='agent').first()

    # Resolve user for order (from customer record if available)
    order_user = None
    if customer_id:
        from .models import Customer as CustomerModel
        cust = CustomerModel.objects.filter(pk=customer_id).select_related('user').first()
        if cust and cust.user:
            order_user = cust.user

    # If linked to existing order, update its payment status instead of creating new
    created_order = None
    if order_id:
        existing_order = Order.objects.filter(pk=order_id).first()
        if existing_order:
            from django.db.models import Sum
            total_paid_from_billings = 0
            for billing in Billing.objects.filter(created_at__gte=existing_order.created_at):
                billing_product_ids = set(billing.items.values_list('product_id', flat=True))
                order_product_ids = set(existing_order.items.values_list('product_id', flat=True))
                if billing_product_ids & order_product_ids:
                    total_paid_from_billings += float(billing.amount_paid)
            total_paid_from_payments = existing_order.payments.aggregate(total=Sum('amount'))['total'] or 0
            total_paid_combined = float(total_paid_from_billings) + float(total_paid_from_payments)
            order_total = float(existing_order.total)
            if total_paid_combined >= order_total:
                existing_order.payment_status = 'paid'
            elif total_paid_combined > 0:
                existing_order.payment_status = 'partial'
            else:
                existing_order.payment_status = 'unpaid'
            existing_order.save()
            created_order = existing_order
    else:
        # Create a new Order with status=pending
        new_order = Order.objects.create(
            user=order_user,
            status='pending',
            delivery_type=order_delivery_type,
            full_name=order_full_name,
            phone=order_phone,
            email=order_email,
            address=order_address,
            city=order_city,
            billing_address=billing_address,
            billing_city=billing_city,
            billing_contact=billing_contact,
            billing_org_name=billing_org_name,
            billing_person_name=billing_person_name,
            pan_number=pan_number,
            payment_method=order_payment_method,
            payment_status=order_pay_status,
            referred_agent=referred_agent,
            subtotal=subtotal,
            delivery_charge=0,
            total=total,
        )
        for item in bill_items:
            OrderItem.objects.create(
                order=new_order,
                product=item['product'],
                product_name=item['product'].name,
                product_sku=item['product'].sku,
                unit_price=item['price'],
                quantity=item['qty'],
            )
        # Record payment if any amount was paid
        if total_paid > 0:
            OrderPayment.objects.create(
                order=new_order,
                amount=total_paid,
                payment_method='cash' if payment_method == 'cash' else ('card' if payment_method == 'card' else 'upi'),
                note=f'POS Bill #{bill.bill_number}',
                recorded_by=request.user,
            )
        created_order = new_order
    
    return JsonResponse({
        'bill_number': bill.bill_number,
        'bill_id': bill.pk,
        'order_id': created_order.pk if created_order else None,
        'order_number': created_order.order_number if created_order else None,
    })




# -- Package Management --------------------------------------------------------

@login_required(login_url='panel_login')
@permission_required('packages', 'view')
def panel_packages(request):
    packages = Package.objects.all().order_by('-created_at')
    return render(request, 'panel/packages.html', {'packages': packages})


def _handle_package_images(request, package):
    
    # 1. Delete marked images
    delete_ids = request.POST.getlist('delete_images')
    if delete_ids:
        images_to_delete = PackageImage.objects.filter(pk__in=delete_ids, package=package)
        for img in images_to_delete:
            img.image.delete(save=False)  # Delete file from storage
            img.delete()

    # 2. Set primary image from radio selection
    primary_id = request.POST.get('primary_image')
    if primary_id:
        PackageImage.objects.filter(package=package).update(is_primary=False)
        PackageImage.objects.filter(pk=primary_id, package=package).update(is_primary=True)

    # 3. Upload new images (max 4 total)
    uploaded_files = request.FILES.getlist('images')
    if uploaded_files:
        current_count = package.images.count()
        slots_available = 4 - current_count
        files_to_upload = uploaded_files[:slots_available]  # Respect max 4 limit

        has_primary = package.images.filter(is_primary=True).exists()

        for i, image_file in enumerate(files_to_upload):
            is_primary = (not has_primary and i == 0)  # First upload becomes primary if none exists
            PackageImage.objects.create(
                package=package,
                image=image_file,
                is_primary=is_primary,
                order=current_count + i
            )
            if is_primary:
                has_primary = True


@login_required(login_url='panel_login')
@permission_required('packages', 'create')
def panel_package_add(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        sku = request.POST.get('sku')
        description = request.POST.get('description', '')
        overall_discount = request.POST.get('overall_discount', 0)
        selling_price = request.POST.get('selling_price')

        package = Package.objects.create(
            name=name, sku=sku, description=description,
            overall_discount=overall_discount, selling_price=selling_price
        )

        # Add items
        product_ids = request.POST.getlist('product_id[]')
        quantities = request.POST.getlist('quantity[]')
        item_discounts = request.POST.getlist('item_discount[]')

        for pid, qty, disc in zip(product_ids, quantities, item_discounts):
            if pid and qty:
                PackageItem.objects.create(
                    package=package,
                    product_id=int(pid),
                    quantity=int(qty),
                    item_discount=float(disc or 0)
                )

        # ? Handle images
        _handle_package_images(request, package)

        messages.success(request, f'Package "{name}" created successfully!')
        return redirect('panel_packages')

    products = Product.objects.filter(is_active=True).order_by('name')
    return render(request, 'panel/package_form.html', {'products': products})


@login_required(login_url='panel_login')
@permission_required('packages', 'edit')
def panel_package_edit(request, pk):
    package = get_object_or_404(Package, pk=pk)

    if request.method == 'POST':
        package.name = request.POST.get('name')
        package.sku = request.POST.get('sku')
        package.description = request.POST.get('description', '')
        package.overall_discount = request.POST.get('overall_discount', 0)
        package.selling_price = request.POST.get('selling_price')
        package.is_active = request.POST.get('is_active') == 'on'
        package.save()

        # Replace all items
        package.items.all().delete()
        product_ids = request.POST.getlist('product_id[]')
        quantities = request.POST.getlist('quantity[]')
        item_discounts = request.POST.getlist('item_discount[]')

        for pid, qty, disc in zip(product_ids, quantities, item_discounts):
            if pid and qty:
                PackageItem.objects.create(
                    package=package,
                    product_id=int(pid),
                    quantity=int(qty),
                    item_discount=float(disc or 0)
                )

        # ? Handle images
        _handle_package_images(request, package)

        messages.success(request, f'Package "{package.name}" updated successfully!')
        return redirect('panel_packages')

    products = Product.objects.filter(is_active=True).order_by('name')
    return render(request, 'panel/package_form.html', {'package': package, 'products': products})

    
@login_required(login_url='panel_login')
@permission_required('packages', 'delete')
def panel_package_delete(request, pk):
    package = get_object_or_404(Package, pk=pk)
    package.delete()
    messages.success(request, 'Package deleted successfully!')
    return redirect('panel_packages')










